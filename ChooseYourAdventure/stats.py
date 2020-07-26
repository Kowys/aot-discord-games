import discord
from ChooseYourAdventure import book
from ChooseYourAdventure import book2
from openpyxl import load_workbook
import math

class Affinities():
    def __init__(self, book):
        if book == 1:
            self.eren = 0
            self.mikasa = 0
            self.armin = 0
            self.jean = 0
            self.krista = 0
            self.sasha = 0
        elif book == 2:
            self.eren = 0
            self.mikasa = 0
            self.armin = 0
            self.jean = 0
            self.krista = 0
            self.sasha = 0
            self.hanji = 0
            self.levi = 0
            self.annie = 0

class Statuses():
    def __init__(self, book):
        if book == 1:
            self.eren = 'Alive'
            self.mikasa = 'Alive'
            self.armin = 'Alive'
            self.jean = 'Alive'
            self.krista = 'Alive'
            self.sasha = 'Alive'
            self.marco = 'Alive'
            self.franz = 'Alive'
            self.hannah = 'Alive'
            self.squad34 = 'Alive'
            self.levi = 'Alive'
            self.rico = 'Alive'
            self.ian = 'Alive'
            self.mitabi = 'Alive'
        elif book == 2:
            self.eren = 'Alive'
            self.mikasa = 'Alive'
            self.armin = 'Alive'
            self.jean = 'Alive'
            self.krista = 'Alive'
            self.sasha = 'Alive'
            self.levi = 'Alive'
            self.hanji = 'Alive'
            self.annie = 'Alive'
            self.petra = 'Alive'
            self.oluo = 'Alive'
            self.gunther = 'Alive'
            self.eld = 'Alive'
            self.dieter = 'Alive'
            self.lukesiss = 'Alive'

class Flags():
    def __init__(self, book):
        if book == 1:
            self.flagD = 'No'
            self.flagL = 'No'
            self.levihelp = 'No'
            self.witheren = 'No'
            self.sashaeaten = 'No'
            self.armindressed = 'No'
            self.armintutorial = 'No'
            self.memelevel = 0
        elif book == 2:
            self.keyA = 'No'
            self.keyY = 'No'
            self.keyR = 'No'
            self.keyS = 'No'
            self.killcount = 0

class State():
    def __init__(self):
        # Initialize as Book 1
        self.affinities = Affinities(1)
        self.statuses = Statuses(1)
        self.flags = Flags(1)
        self.book = book.Book(self)
        self.book_no = 1
        self.cur_player_id = None
        self.select_book = False
        self.timer = 0
        self.game_channel = None

    def reset_game(self):
        self.affinities = Affinities(1)
        self.statuses = Statuses(1)
        self.flags = Flags(1)
        self.book = book.Book(self)
        self.book_no = 1
        self.cur_player_id = None
        self.select_book = False
        self.timer = 0
        self.game_channel = None
        
    
    def check_book(self, player):
        wb = load_workbook("ChooseYourAdventure/player data.xlsx")
        player_data = wb['Endings']
        my_endings = []
        for line in player_data:
            if str(player.id) == line[0].value and line[1].value not in my_endings:
                my_endings.append(line[1].value)

        good_endings = ['An Ordinary Moment of Happiness', 'Jean Kirstein of the Survey Corps', 'A Narrow Victory', 'Armin Arlert\'s Dream', 'Captain Levi\'s Recruit', 'Mikasa\'s True Face', 'Nameless Hero', 
        'Eren Yeager\'s Hand', 'Sasha Blouse\'s Promise', 'The Girl Who Hid Her True Self', 'No Regrets', 'Jean of the Military Police', 'Joining the Garrison']

        for ending in my_endings:
            if ending in good_endings:
                return True

        return False

    def switch_book(self):
        self.affinities = Affinities(2)
        self.statuses = Statuses(2)
        self.flags = Flags(2)
        self.book = book2.Book(self)
        self.book_no = 2

    def get_annie_secrets(self, user_obj):
        wb = load_workbook("ChooseYourAdventure/player data.xlsx")
        annie_data = wb['Annie progress']
        my_secrets = ['one']
        cur_secret = self.book.annie_secrets[self.book.cur_page[0]]
        for line in annie_data:
            if str(user_obj.id) == line[0].value and (line[1].value not in my_secrets):
                my_secrets.append(line[1].value)     

        if cur_secret not in my_secrets:
            return 'Nice try, ' + user_obj.mention + ', but you do not have access to this secret yet!'
        else:
            # Record secret
            annie_data.append([str(user_obj.id), self.book.annie_secrets2[self.book.cur_page[0]]])
            wb.save("ChooseYourAdventure/player data.xlsx")

            # Go to secret page and return content
            msg_content = self.book.page_flipper('s')
            return msg_content

    def log_ending(self, author):
        wb = load_workbook("ChooseYourAdventure/player data.xlsx")

        # Add ending to game statistics
        game_data = wb['Global']
        ending = self.book.endings[self.book.cur_page[0]]
        ending_map = {'Trial of Eren and Mikasa': 15, 'A Soldier\'s Duty': 16, 'An Ordinary Moment of Happiness': 3, 'Failure of the Reclamation Plan': 17, 'Jean Kirstein of the Survey Corps': 4, 
        'A Narrow Victory': 5, 'Armin Arlert\'s Dream': 6, 'A Regular Soldier': 18, '104th Annihilated at HQ': 19, 'Junior High': 25, 'The Fall of Wall Rose': 20, 
        'Failure to Reclaim Trost District': 21, 'A Moment\'s Peace': 22, 'Eren Flees': 23, 'Captain Levi\'s Recruit': 7, 'Mikasa\'s True Face': 8, 'Nameless Hero': 9, 'Eren Yeager\'s Hand': 10, 
        'Sasha Blouse\'s Promise': 11, 'The Girl Who Hid Her True Self': 12, 'The Death of a Merchant': 24, 'No Regrets': 13, 'Jean of the Military Police': 14, 'Joining the Garrison': 2}

        cell = 'B' + str(ending_map[ending])
        game_data[cell] = game_data[cell].value + 1
        game_data['B1'] = game_data['B1'].value + 1
        wb.save("ChooseYourAdventure/player data.xlsx")

        # Logs the ending the player achieved
        player_data = wb['Endings']
        former_endings = []
        for line in player_data:
            if str(author.id) == line[0].value and ([line[0].value, line[1].value] not in former_endings):
                former_endings.append([line[0].value, line[1].value])
        former_ending_count = len(former_endings)

        # Add ending the player got
        player_data.append([str(author.id), self.book.endings[self.book.cur_page[0]]])
        wb.save("ChooseYourAdventure/player data.xlsx")

        # Inform player of rank ups
        my_endings = []
        for line in player_data:
            if str(author.id) == line[0].value and ([line[0].value, line[1].value] not in my_endings):
                my_endings.append([line[0].value, line[1].value])
        ending_count = len(my_endings)
        rank_ups = {1: ['Trainee', 'Recruit'], 2: ['Recruit', 'Private'], 3: ['Private', 'Soldier'], 5: ['Soldier', 'Elite Soldier'], 8: ['Elite Soldier', 'Officer'], 10: ['Officer', 'Team Leader'], 
        13: ['Team Leader', 'Senior Team Leader'], 15: ['Senior Team Leader', 'Squad Leader'], 20: ['Squad Leader', 'Squad Captain'], 24: ['Squad Captain', 'Commander']}
        
        if (ending_count > former_ending_count) and (ending_count in rank_ups):
            new_rank = rank_ups[ending_count][1]
            former_rank = rank_ups[ending_count][0]

            # Rank-up message
            promotion_info = discord.Embed(title='🎖' + author.name + ' has been promoted to ' + new_rank + '!🎖', colour=0xE5D2BB)
            promotion_info.set_thumbnail(url = author.avatar_url)
            promotion_info.add_field(name = 'Congratulations!', value =  '**' + former_rank + '** ➡ **' + new_rank + '**')

            return promotion_info

        # If no promotion
        else:
            return None

    def get_stats(self):
        stats_info = discord.Embed(title='Battle information\n', colour=0xE5D2BB)

        if self.book_no == 1:
            stats_info.add_field(name = '**Statuses**', value = 'Eren: ' + self.statuses.eren + \
            '\nMikasa: ' + self.statuses.mikasa + \
            '\nArmin: ' + self.statuses.armin + \
            '\nJean: ' + self.statuses.jean + \
            '\nKrista: ' + self.statuses.krista + \
            '\nSasha: ' + self.statuses.sasha + \
            '\nMarco: ' + self.statuses.marco + \
            '\nFranz: ' + self.statuses.franz + \
            '\nHannah: ' + self.statuses.hannah + \
            '\nSquad 34: ' + self.statuses.squad34 + \
            '\nLevi: ' + self.statuses.levi + \
            '\nRico: ' + self.statuses.rico + \
            '\nIan: ' + self.statuses.ian + \
            '\nMitabi: ' + self.statuses.mitabi) 

            stats_info.add_field(name = '**Affinities**', value = 'Eren: ' + str(self.affinities.eren) + \
            '\nMikasa: ' + str(self.affinities.mikasa) + \
            '\nArmin: ' + str(self.affinities.armin) + \
            '\nJean: ' + str(self.affinities.jean) + \
            '\nKrista: ' + str(self.affinities.krista) + \
            '\nSasha: ' + str(self.affinities.sasha))
            
            stats_info.add_field(name = '**Flags**', value = 'Flag D : ' + self.flags.flagD + \
            '\nFlag L: ' + self.flags.flagL)

        elif self.book_no == 2:
            stats_info.add_field(name = '**Statuses**', value = 'Eren: ' + self.statuses.eren + \
            '\nMikasa: ' + self.statuses.mikasa + \
            '\nArmin: ' + self.statuses.armin + \
            '\nJean: ' + self.statuses.jean + \
            '\nKrista: ' + self.statuses.krista + \
            '\nSasha: ' + self.statuses.sasha + \
            '\nLevi: ' + self.statuses.levi + \
            '\nHanji: ' + self.statuses.hanji + \
            '\nAnnie: ' + self.statuses.annie + \
            '\nPetra: ' + self.statuses.petra + \
            '\nOluo: ' + self.statuses.oluo + \
            '\nGunther: ' + self.statuses.gunther + \
            '\nEld: ' + self.statuses.eld + \
            '\nDieter Ness: ' + self.statuses.dieter + \
            '\nLuke Siss: ' + self.statuses.lukesiss) 

            stats_info.add_field(name = '**Affinities**', value = 'Eren: ' + str(self.affinities.eren) + \
            '\nMikasa: ' + str(self.affinities.mikasa) + \
            '\nArmin: ' + str(self.affinities.armin) + \
            '\nJean: ' + str(self.affinities.jean) + \
            '\nKrista: ' + str(self.affinities.krista) + \
            '\nSasha: ' + str(self.affinities.sasha) + \
            '\nHanji: ' + str(self.affinities.hanji) + \
            '\nLevi: ' + str(self.affinities.levi) + \
            '\nAnnie: ' + str(self.affinities.annie))

            keys = ''
            if self.flags.keyA == 'Yes':
                keys += 'Key A\n'
            if self.flags.keyY == 'Yes':
                keys += 'Key Y\n'
            if self.flags.keyR == 'Yes':
                keys += 'Key R\n'
            if self.flags.keyS == 'Yes':
                keys += 'Key S\n'

            stats_info.add_field(name = '**Keys**', value = keys if keys else '-')

            stats_info.add_field(name = '**Kill Count**', value = str(self.flags.killcount))

        return stats_info

    def get_profile(self, user_obj):
        # Retrieve obtained endings from excel sheet
        wb = load_workbook("ChooseYourAdventure/player data.xlsx")
        player_data = wb['Endings']
        my_endings = []
        for line in player_data:
            if str(user_obj.id) == line[0].value and line[1].value not in my_endings:
                my_endings.append(line[1].value)

        ending_count = len(my_endings)
        # if user_obj.id == '238808836075421697':
        #     my_rank = '**Commander 👑**'
        if ending_count == 0:
            my_rank = '**Trainee 🔸**'
        elif ending_count == 1:
            my_rank = '**Recruit 🔸🔸**'
        elif ending_count == 2:
            my_rank = '**Private 🔸🔸🔸**'
        elif ending_count < 5:
            my_rank = '**Soldier 🔹**'
        elif ending_count < 8:
            my_rank = '**Elite Soldier 🔹🔹**'
        elif ending_count < 10:
            my_rank = '**Officer 🔹🔹🔹**'
        elif ending_count < 13:
            my_rank = '**Team Leader 💠**'
        elif ending_count < 15:
            my_rank = '**Senior Team Leader 💠🔹**'
        elif ending_count < 20:
            my_rank = '**Squad Leader ⭐**'
        elif ending_count < 24:
            my_rank = '**Squad Captain 🌟**'
        elif ending_count == 24:
            my_rank = '**Commander 👑**'
        desc = 'Rank: ' + my_rank + '\n' + 'Number of endings obtained: ' + str(ending_count) + '/24'
        progress_info = discord.Embed(title='Player information\n', description = desc, colour=0xE5D2BB)
        # Include number of games played, and other interesting stats?

        username, user_url = user_obj.name, user_obj.avatar_url
        progress_info.set_author(name = username, icon_url = user_url)
        progress_info.set_thumbnail(url = user_url)

        endings = [['An Ordinary Moment of Happiness', 'Jean Kirstein of the Survey Corps', 'A Narrow Victory', 'Armin Arlert\'s Dream', 'Captain Levi\'s Recruit', 'Mikasa\'s True Face', 'Nameless Hero', 
        'Eren Yeager\'s Hand', 'Sasha Blouse\'s Promise', 'The Girl Who Hid Her True Self', 'No Regrets', 'Jean of the Military Police', 'Joining the Garrison'], 
        ['Trial of Eren and Mikasa', 'A Soldier\'s Duty', 'Failure of the Reclamation Plan', 'A Regular Soldier', '104th Annihilated at HQ', 'The Fall of Wall Rose', 'Failure to Reclaim Trost District',
        'A Moment\'s Peace', 'Eren Flees', 'The Death of a Merchant'],
        ['Junior High']]

        good_endings, bad_endings, secret_endings = '', '', ''
        for ending in my_endings:
            if ending in endings[0]:
                good_endings += ending + '\n'
            elif ending in endings[1]:
                bad_endings += ending + '\n'
            elif ending in endings[2]:
                secret_endings += ending + '\n'

        progress_info.add_field(name = '**Good Endings**', value = '-' if good_endings == '' else good_endings) 
        progress_info.add_field(name = '**Bad Endings**', value = '-' if bad_endings == '' else bad_endings)
        progress_info.add_field(name = '**Secret Endings**', value = '-' if secret_endings == '' else secret_endings, inline = False)

        return progress_info

    def get_leaderboard(self, server, page=1):
        wb = load_workbook("ChooseYourAdventure/player data.xlsx")
        player_data = wb['Endings']
        # {playerid1: [end1, end2, ...], playerid2: [end1, end2, ...], ...}
        player_endings = {}
        ending_no = 0
        for line in player_data:
            ending_no += 1
            if line[0].value in player_endings:
                if line[1].value not in [x[0] for x in player_endings[line[0].value]]:
                    # New ending, add to all endings, including when the new ending is obtained
                    player_endings[line[0].value].append([line[1].value, ending_no])
            else:
                player_endings[line[0].value] = [[line[1].value, ending_no]]

        server_users = server.members
        server_players = []
        for user in server_users:
            if str(user.id) in player_endings:
                server_players.append([user.name, len(player_endings[str(user.id)]), player_endings[str(user.id)][-1][1]])
        
        # Sort according to when player obtained their last ending (earlier = smaller number)
        server_players.sort(key = lambda x: x[2])
        # Sort according to ending count
        server_players.sort(key = lambda x: x[1], reverse = True)

        # Top 10x players
        try: 
            page_no = int(page)
        except:
            page_no = 1

        if page_no < 1:
            page_no = 1

        # Normalize to number of server users
        page_no = min(math.ceil(len(server_players) / 10), page_no)

        leaderboard = discord.Embed(title = '🏆 Leaderboard for Choose Your Adventure 🏆', description = 'Page ' + str(page_no), colour=0xE5D2BB)

        for rank in range((page_no - 1) * 10 + 1, min(len(server_players) + 1, page_no * 10 + 1)):
            ending_count = server_players[rank-1][1]
            if ending_count == 0:
                insignia = '🔸'
            elif ending_count == 1:
                insignia = '🔸🔸'
            elif ending_count == 2:
                insignia = '🔸🔸🔸'
            elif ending_count < 5:
                insignia = '🔹'
            elif ending_count < 8:
                insignia = '🔹🔹'
            elif ending_count < 10:
                insignia = '🔹🔹🔹'
            elif ending_count < 13:
                insignia = '💠'
            elif ending_count < 15:
                insignia = '💠🔹'
            elif ending_count < 20:
                insignia = '⭐'
            elif ending_count < 24:
                insignia = '🌟'
            elif ending_count == 24:
                insignia = '👑'
            leaderboard.add_field(name = insignia + ' | #' + str(rank) + '  ' + server_players[rank-1][0], 
                                  value = '# Endings: ￵**' + str(server_players[rank-1][1]) + '**', inline = False)

        return leaderboard

    def get_gamestats(self):
        # Returns overall game statistics
        wb = load_workbook("ChooseYourAdventure/player data.xlsx")
        game_data = wb['Global']
        game_stats = discord.Embed(title = '🔦 Choose Your Adventure Stats', colour=0xE5D2BB)
        game_stats.add_field(name = 'Total Number Of Endings Obtained', value = '**' + str(game_data['B1'].value) + '**', inline = False)

        endings_list = []
        i = 0
        for row in game_data:
            i += 1
            if i > 1:
                endings_list.append([game_data['A' + str(i)].value, game_data['B' + str(i)].value])
            if i >= 25:
                break

        endings_list.sort(key = lambda x: x[1], reverse = True)

        all_endings = ''
        for ending in endings_list:
            all_endings += ending[0] + ': ' + str(ending[1]) + '\n'

        game_stats.add_field(name = 'Endings', value = all_endings)
        return game_stats

    def get_commands(self):
        # Returns the list of commands
        commands_list = discord.Embed(title = 'List of commands for Choose Your Adventure', colour=0xE5D2BB)
        commands_list.add_field(name = '~start', value = 'Starts a new game.')
        commands_list.add_field(name = '~intro', value = 'Gives a short introduction to the game.')
        commands_list.add_field(name = '~reset', value = 'Resets the game.')
        commands_list.add_field(name = '~stats', value = 'Brings up the Battle Report Card during a game.')
        commands_list.add_field(name = '~profile <@person>', value = 'Checks the profile of a given user. Use just ~profile to check your own profile.')
        commands_list.add_field(name = '~leaderboard/~lb', value = 'Shows the leaderboard, listing the top 10 players on the server. Add a number behind to see subsequent pages (e.g. ~lb 2).')
        commands_list.add_field(name = '~gamestats', value = 'Brings up the overall game statistics.')
        return commands_list