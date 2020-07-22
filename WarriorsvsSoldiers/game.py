import discord
import random
from openpyxl import load_workbook
import math

class State():
    def __init__(self):
        # Constants
        self.intro_msg = 'Welcome to ⚔**Warriors vs Soldiers**🛡!\n\n\
In this Attack on Titan themed game, the Soldiers of the Survey Corps embark on expeditions to reach the Basement, which they believe \
holds the secrets about the world. However, the Titan shifters known as Warriors have infiltrated their ranks, and will do anything it takes to sabotage the expeditions and destroy the Walls.\n\n\
As a participant in this game, your abilities of logical deduction, deception, and teamwork will be put to the ultimate test.\n\n\
Will the Soldiers reach the Basement and find out the truth about the world? \
Or will the Warriors destroy the Walls and wipe out humanity? You decide!\n\n\
❗Type **`~host`** to begin a new game.\n\n\
❗You may type **`~help`** at any time to see the list of commands.'

        self.reset_msg = discord.Embed(title = 'Game Reset!', description = 'Type **`~host`** to create a new lobby.\nType **`~intro`** for more information about the game.', colour=0x0013B4)

        self.start_msg = 'Assigning roles...'
        self.roles_assigned_msg = 'Roles assigned!\n\nWhen everyone is ready, the host may type **~next** to begin the first expedition.'

        self.player_roles = {5: ['soldier','soldier','warrior','warrior','coordinate'], 
                             6: ['soldier','soldier','soldier','warrior','warrior','coordinate'],
                             7: ['soldier','soldier','soldier','warrior','warrior','warrior','coordinate'], 
                             8: ['soldier','soldier','soldier','soldier','warrior','warrior','warrior','coordinate'], 
                             9: ['soldier','soldier','soldier','soldier','soldier','warrior','warrior','warrior','coordinate'], 
                             10: ['soldier','soldier','soldier','soldier','soldier','warrior','warrior','warrior','warrior','coordinate']}
        self.warrior_roles = ['warrior','warchief','false king','ymir','spy']
        self.expedition_sizes = {5:[2,3,2,3,3], 6:[2,3,4,3,4], 7:[2,3,3,4,4], 8:[3,3,4,4,5], 9:[3,4,4,5,5], 10:[3,3,4,5,5]}
        self.expedition_advantages = {5:-50, 6:50, 7:-100, 8:0, 9:100, 10:-150}
        self.roles_advantages = {'queen': {5:150, 6:130, 7:120, 8:100, 9:80, 10:70},
                                 'ackerman': {5:100, 6:120, 7:150, 8:160, 9:180, 10:200},
                                 'mike': {5:160, 6:140, 7:120, 8:100, 9:90, 10:80},
                                 'scout': {5:-120, 6:-90, 7:-80, 8:-60, 9:-40, 10:-40},
                                 'warchief': {5:-120, 6:-110, 7:-90, 8:-80, 9:-60, 10:-50},
                                 'false king': {5:-140, 6:-120, 7:-100, 8:-80, 9:-60, 10:-40},
                                 'ymir': {5:100, 6:90, 7:60, 8:50, 9:50, 10:30},
                                 'spy': {5:-140, 6:-160, 7:-180, 8:-200, 9:-220, 10:-240}}
        self.blessing_advantages = {5:150, 6:120, 7:100, 8:80, 9:60, 10:50}


        self.number_games_achievements = {10:'Bronze', 25:'Silver', 50:'Gold', 100:'Platinum', 250:'Diamond', 500:'Master', 1000:'Grandmaster'}
        self.consecutive_wins_achievements = {2:'Bronze', 3:'Silver', 5:'Gold', 7:'Platinum', 10:'Diamond', 15:'Master', 20:'Grandmaster'}
        self.soldier_wins_achievements = {3:'Bronze', 8:'Silver', 15:'Gold', 25:'Platinum', 50:'Diamond', 100:'Master', 200:'Grandmaster'}
        self.warrior_wins_achievements = {3:'Bronze', 8:'Silver', 15:'Gold', 25:'Platinum', 50:'Diamond', 100:'Master', 200:'Grandmaster'}
        self.coordinate_wins_achievements = {3:'Bronze', 8:'Silver', 15:'Gold', 25:'Platinum', 50:'Diamond', 100:'Master', 200:'Grandmaster'}
        self.queen_wins_achievements = {3:'Bronze', 8:'Silver', 15:'Gold', 25:'Platinum', 50:'Diamond', 100:'Master', 200:'Grandmaster'}
        self.warchief_wins_achievements = {3:'Bronze', 8:'Silver', 15:'Gold', 25:'Platinum', 50:'Diamond', 100:'Master', 200:'Grandmaster'}

        self.achievement_rewards = {'Bronze': 20, 'Silver': 40, 'Gold': 60, 'Platinum': 80, 'Diamond': 100, 'Master': 150, 'Grandmaster': 250}
        self.badge_emojis = {'Bronze': '🥉', 'Silver': '🥈', 'Gold': '🥇', 'Platinum': '💠', 'Diamond': '💎', 'Master': '👑', 'Grandmaster': '🎓'}

        # Variables
        self.status = 'waiting for game' # waiting for players, assigning roles, expedition selection, expedition approval, expedition decision, expedition over, choose coordinate, game ended

        self.game_channel = None
        self.server = None
        self.game_host = None
        self.gametype = 'Ranked'
        self.players = [] # each player is a list of two elements: the player object and the player's role (str)
        self.message_box = []
        self.newroles = []
        self.ymir_blessing = False
        self.blessed = []
        self.currently_blessed = None
        self.randomroles = False
        self.paths = False
        self.paths_holders = []
        
        self.cur_expedition = 0
        self.expedition_squad = []
        self.expedition_approval = []
        self.votes_flipped = False
        self.flipping_votes = False
        self.expedition_result = []

        self.expedition_history = []
        self.successful_expeditions = 0
        self.walls_destroyed = 0
        self.wall_secured = False

    def host(self, player, server, ranked):
        self.server = server
        wb = load_workbook("WarriorsvsSoldiers/database.xlsx")
        player_data = wb['Player records']
        wb = load_workbook("WarriorsvsSoldiers/blacklist.xlsx")
        blacklist_players = wb['Blacklist']
        if str(player.id) in [row[0].value for row in blacklist_players]:
            return 'You have been blacklisted from the game. Shame on you...'
           
        if self.status == 'waiting for players':
            return '**' + self.game_host.name + '** is already the host!'
        elif self.status == 'waiting for game' or self.status == 'game ended':
            if ranked == False:
                self.gametype = 'Casual'
            self.game_host = player
            self.status = 'waiting for players'
            self.players = [[player, None]]
            self.newroles = []
            self.ymir_blessing = False
            self.blessed = []
            self.currently_blessed = None
            self.randomroles = False
            self.paths = False
            self.paths_holders = []

            self.cur_expedition = 0
            self.expedition_squad = []
            self.expedition_approval = []
            self.votes_flipped = False
            self.flipping_votes = False
            self.expedition_result = []

            self.expedition_history = []
            self.successful_expeditions = 0
            self.walls_destroyed = 0
            self.wall_secured = False

            msg = '**' + player.name + '** has started a new lobby! Type `~join` to join the game!'

            # Get total players in server
            player_rankings = [] # [[player id, sr], [...], ...]
            # Put all players into a list
            for row in player_data:
                if row[0].value == 'ID':
                    continue
                else:
                    player_rankings.append([row[0].value, row[1].value])
            
            # Sort by Exp from biggest to smallest
            player_rankings.sort(key = lambda x: x[1], reverse = True)

            # Get players in server
            server_users = list(map(lambda y: str(y.id), server.members))
            server_players = list(filter(lambda x:x[0] in server_users, player_rankings))

            for person in server_players:
                if person[0] == str(player.id):
                    msg = '👑 | Humanity\'s strongest soldier ' + msg
                break

            return msg
        else:
            return 'The game has already started!'

    def join(self, player, server):
        wb = load_workbook("WarriorsvsSoldiers/database.xlsx")
        player_data = wb['Player records']
        wb = load_workbook("WarriorsvsSoldiers/blacklist.xlsx")
        blacklist_players = wb['Blacklist']
        if str(player.id) in [row[0].value for row in blacklist_players]:
            return 'You have been blacklisted from the game. Shame on you...'

        elif self.status == 'waiting for players':
            if player in list(map(lambda x:x[0], self.players)):
                return 'You are already in the lobby!'
            elif len(self.players) >= 10:
                return 'Lobby is full! (Maximum of 10 players)'
            else:
                self.players.append([player, None])
                msg = '**' + player.name + '** has joined the game!'
                # Get total players in server
                player_rankings = [] # [[player id, sr], [...], ...]
                # Put all players into a list
                for row in player_data:
                    if row[0].value == 'ID':
                        continue
                    else:
                        player_rankings.append([row[0].value, row[1].value])
                
                # Sort by Exp from biggest to smallest
                player_rankings.sort(key = lambda x: x[1], reverse = True)

                # Get players in server
                server_users = list(map(lambda y: str(y.id), server.members))
                server_players = list(filter(lambda x:x[0] in server_users, player_rankings))

                for person in server_players:
                    if person[0] == str(player.id):
                        msg = '👑 | Humanity\'s strongest soldier ' + msg
                    break

                return msg

        elif self.status == 'waiting for game' or self.status.startswith('game ended'):
            return 'There is no open lobby at the moment!'
        else:
            return 'The game has already started!'

    def leave(self, player):
        if self.status == 'waiting for players':
            if player not in list(map(lambda x:x[0], self.players)):
                return 'You are not in the lobby!'
            # Resets if no one is left
            elif len(self.players) == 1:
                self.reset(player)
                return '**' + player.name + '** has left the lobby! Lobby closed.'
            # Also resets if host leaves
            elif player == self.game_host:
                self.reset(player)
                return 'The host, **' + player.name + '**, has left the lobby! Lobby closed.'
            else:
                self.players.remove([player, None])
                return '**' + player.name + '** has left the lobby!'
        elif self.status == 'waiting for game' or self.status.startswith('game ended'):
            return 'There is no open lobby at the moment!'
        else:
            return 'The game has already started!'

    def kick(self, author, player):
        if self.status == 'waiting for players':
            if author != self.game_host and author.id != 238808836075421697:
                return 'Only the host can kick other players!'
            elif player not in list(map(lambda x:x[0], self.players)):
                return '**' + player.name + '** is not in the lobby!'
            elif player == self.game_host:
                return 'You can\'t kick yourself!'
            else:
                self.players.remove([player, None])
                return '**' + player.name + '** has been removed from the lobby!'
        elif self.status == 'waiting for game':
            return 'There is no open lobby at the moment!'
        else:
            return 'The game has already started!'

    def ban(self, player):
        wb = load_workbook("WarriorsvsSoldiers/blacklist.xlsx")
        blacklist_players = wb['Blacklist']
        if str(player.id) in [row[0].value for row in blacklist_players]:
            return '**' + player.name + '** has already been blacklisted from the game!'
        else:
            blacklist_players.append([str(player.id)])
            wb.save("WarriorsvsSoldiers/blacklist.xlsx")
            return '**' + player.name + '** has been blacklisted from the game!'

    def unban(self, player):
        wb = load_workbook("WarriorsvsSoldiers/blacklist.xlsx")
        blacklist_players = wb['Blacklist']
        if str(player.id) not in [row[0].value for row in blacklist_players]:
            return '**' + player.name + '** has not been blacklisted from the game!'
        else:
            i = 0
            for row in blacklist_players:
                i += 1
                if str(player.id) == row[0].value:
                    blacklist_players['A'+ str(i)] = 0
            wb.save("WarriorsvsSoldiers/blacklist.xlsx")
            return '**' + player.name + '** is no longer blacklisted from the game!'

    def add_with_role_count_check(self, role):
        soldier_dict = {'queen': '👼**Queen**👼',
                        'ackerman': '💂**Ackerman**💂',
                        'mike': '<:aotSmirk2:455223780957224961>**Mike Zacharias** <:aotSmirk2:455223780957224961>',
                        'scout': '🏇**Scout**🏇'}

        warrior_dict = {'warchief': '🦹‍♂️**Warchief**🦹‍♂️',
                        'ymir': '🤷‍♀️**Ymir**🤷‍♀️',
                        'false king': '🕴**False King**🕴',
                        'spy': '🕵️‍♀️**Spy**🕵️‍♀️'}
        
        if role in soldier_dict:
            soldier_role_count = len(list(filter(lambda role:role not in self.warrior_roles, self.newroles)))
            if soldier_role_count >= 3:
                if len(self.players) >= 8:
                    self.newroles.append(role)
                    return 'The role {} has been added to the game!'.format(soldier_dict[role])
                else:
                    return 'At least 8 players are needed to add another optional soldier role!'
            elif soldier_role_count >= 2:
                if len(self.players) >= 6:
                    self.newroles.append(role)
                    return 'The role {} has been added to the game!'.format(soldier_dict[role])
                else:
                    return 'At least 6 players are needed to add another optional soldier role!'
            else:
                self.newroles.append(role)
                return 'The role {} has been added to the game!'.format(soldier_dict[role])

        elif role in warrior_dict:
            warrior_role_count = len(list(filter(lambda role:role in self.warrior_roles, self.newroles)))
            if warrior_role_count >= 3:
                if len(self.players) >= 10:
                    self.newroles.append(role)
                    return 'The role {} has been added to the game!'.format(warrior_dict[role])
                else:
                    return 'At least 10 players are needed to add another optional warrior role!'
            elif warrior_role_count >= 2:
                if len(self.players) >= 7:
                    self.newroles.append(role)
                    return 'The role {} has been added to the game!'.format(warrior_dict[role])
                else:
                    return 'At least 7 players are needed to add another optional warrior role!'
            else:
                self.newroles.append(role)
                return 'The role {} has been added to the game!'.format(warrior_dict[role])

    def addrole(self, role, player, randomroles=False):
        if self.status == 'waiting for players' or randomroles == True:
            if player == self.game_host or player.id == 238808836075421697:
                if role == 'blessing' or role == 'ymir\'s blessing' or role == 'ymirs blessing' or role == 'ymir blessing':
                    if self.ymir_blessing == False:
                        self.ymir_blessing = True
                        return '🔮**Ymir\'s Blessing**🔮 has been enabled!'
                    else:
                        return '🔮**Ymir\'s Blessing**🔮 has already been enabled!'

                elif role == 'paths':
                    if self.paths == False:
                        self.paths = True
                        return 'The power of the 📢**Paths**📢 has been enabled!'
                    else:
                        return 'The power of the 📢**Paths**📢 has already been enabled!'

                elif role == 'queen':
                    if role not in self.newroles:
                        return self.add_with_role_count_check('queen')
                    else:
                        return 'The 👼**Queen**👼 role is already in the game!'

                elif role == 'ackerman':
                    if role not in self.newroles:
                        return self.add_with_role_count_check('ackerman')
                    else:
                        return 'The 💂**Ackerman**💂 role is already in the game!'

                elif role == 'mike' or role == 'mike zacharias' or role == 'zacharias':
                    role = 'mike'
                    if role not in self.newroles:
                        return self.add_with_role_count_check('mike')
                    else:
                        return '<:aotSmirk2:455223780957224961>**Mike Zacharias** <:aotSmirk2:455223780957224961> is already in the game!'

                elif role == 'scout':
                    if role not in self.newroles:
                        return self.add_with_role_count_check('scout')
                    else:
                        return 'The 🏇**Scout**🏇 role is already in the game!'
                    
                elif role == 'warchief':
                    if role not in self.newroles:
                        return self.add_with_role_count_check('warchief')
                    else:
                        return 'The 🦹‍♂️**Warchief**🦹‍♂️ role is already in the game!'

                elif role == 'ymir':
                    if role not in self.newroles:
                        return self.add_with_role_count_check('ymir')
                    else:
                        return 'The 🤷‍♀️**Ymir**🤷‍♀️ role is already in the game!'

                elif role == 'false king':
                    if role not in self.newroles:
                        if 'queen' not in self.newroles:
                            return 'The 👼**Queen**👼 role needs to be added first!'

                        return self.add_with_role_count_check('false king')
                    else:
                        return 'The 🕴**False King**🕴 role is already in the game!'

                elif role == 'spy':
                    if role not in self.newroles:
                        return self.add_with_role_count_check('spy')
                    else:
                        return 'The 🕵️‍♀️**Spy**🕵️‍♀️ role is already in the game!'

                else:
                    return 'Invalid role!'
            else:
                return 'Only the host can add or remove optional roles!'
        else:
            return 'There is no open lobby at the moment!'

    def removerole(self, role, player):
        if self.status == 'waiting for players':
            if player == self.game_host or player.id == 238808836075421697:
                if role == 'blessing' or role == 'ymir\'s blessing' or role == 'ymirs blessing' or role == 'ymir blessing':
                    if self.ymir_blessing == True:
                        self.ymir_blessing = False
                        return '🔮**Ymir\'s Blessing**🔮 has been disabled!'
                    else:
                        return '🔮**Ymir\'s Blessing**🔮 is not currently enabled!'

                elif role == 'paths':
                    if self.paths == True:
                        self.paths = False
                        return 'The power of the 📢**Paths**📢 has been disabled!'
                    else:
                        return 'The power of the 📢**Paths**📢 is not currently enabled!'

                elif role == 'queen':
                    if role in self.newroles:
                        self.newroles.remove('queen')
                        if 'false king' in self.newroles:
                            self.newroles.remove('false king')
                        return 'The role 👼**Queen**👼 has been removed from the game!'
                    else:
                        return 'The 👼**Queen**👼 role is not in the game!'

                elif role == 'ackerman':
                    if role in self.newroles:
                        self.newroles.remove('ackerman')
                        return 'The role 💂**Ackerman**💂 has been removed from the game!'
                    else:
                        return 'The 💂**Ackerman**💂 role is not in the game!'

                elif role == 'mike' or role == 'mike zacharias' or role == 'zacharias':
                    role = 'mike'
                    if role in self.newroles:
                        self.newroles.remove('mike')
                        return '<:aotSmirk2:455223780957224961>**Mike Zacharias** <:aotSmirk2:455223780957224961> has been removed from the game!'
                    else:
                        return '<:aotSmirk2:455223780957224961>**Mike Zacharias** <:aotSmirk2:455223780957224961> is not in the game!'

                elif role == 'scout':
                    if role in self.newroles:
                        self.newroles.remove('scout')
                        return 'The role 🏇**Scout**🏇 has been removed from the game!'
                    else:
                        return 'The 🏇**Scout**🏇 role is not in the game!'
                    
                elif role == 'warchief':
                    if role in self.newroles:
                        self.newroles.remove('warchief')
                        return 'The role 🦹‍♂️**Warchief**🦹‍♂️ has been removed from the game!'
                    else:
                        return 'The 🦹‍♂️**Warchief**🦹‍♂️ role is not in the game!'

                elif role == 'ymir':
                    if role in self.newroles:
                        self.newroles.remove('ymir')
                        return 'The role 🤷‍♀️**Ymir**🤷‍♀️ has been removed from the game!'
                    else:
                        return 'The 🤷‍♀️**Ymir**🤷‍♀️ role is not in the game!'

                elif role == 'false king':
                    if role in self.newroles:
                        self.newroles.remove('false king')
                        return 'The role 🕴**False King**🕴 has been removed from the game!'
                    else:
                        return 'The 🕴**False King**🕴 role is not in the game!'

                elif role == 'spy':
                    if role in self.newroles:
                        self.newroles.remove('spy')
                        return 'The role 🕵️‍♀️**Spy**🕵️‍♀️ has been removed from the game!'
                    else:
                        return 'The 🕵️‍♀️**Spy**🕵️‍♀️ role is not in the game!'

                else:
                    return 'Invalid role!'
            else:
                return 'Only the host can add or remove optional roles!'
        else:
            return 'There is no open lobby at the moment!'

    def toggle_randomroles(self, player):
        if self.status == 'waiting for players':
            if player == self.game_host or player.id == 238808836075421697:
                if self.randomroles:
                    self.randomroles = False
                    return 'Random roles have been disabled!'
                else:
                    self.randomroles = True
                    return 'Random roles have been enabled!'
            else:
                return 'Only the host can add or remove optional roles!'
        else:
            return 'There is no open lobby at the moment!'

    def randomize_roles(self):
        self.newroles = []
        rolelist = ['queen', 'ackerman', 'mike', 'scout', 'warchief', 'false king', 'ymir', 'spy']
        for role in rolelist:
            to_add = random.randint(0, 1)
            if to_add:
                self.addrole(role, self.game_host, randomroles=True)

    def display_lobby(self):
        players_in_lobby = ''
        for person in self.players:
            if person[0] == self.game_host:
                players_in_lobby = players_in_lobby + '**' + person[0].mention + '** (Host)\n'
            else:
                players_in_lobby = players_in_lobby + '**' + person[0].mention + '**\n'

        lobby = discord.Embed(title = 'Current lobby', colour=0x0013B4)
        lobby.add_field(name = 'Game Type', value = '**' + self.gametype + '**')
        lobby.add_field(name = 'Advantage', value = '⚖️ **Perfectly balanced!** ⚖️' if self.calculate_advantage() == '**None**' else self.calculate_advantage())
        lobby.add_field(name = '**' + str(len(self.players)) + '** player' + ('s' if len(self.players) > 1 else '') + ' in lobby', value = players_in_lobby, inline = False)
        lobby.add_field(name = 'Random Roles', value = '🎲' if self.randomroles else '⬛', inline=False)
        lobby.add_field(name = 'Optional Soldier Roles', value = self.get_newroles('soldier'))
        lobby.add_field(name = 'Optional Warrior Roles', value = self.get_newroles('warrior'))
        lobby.add_field(name = 'In-Game Effects', value = self.get_newroles('game'), inline = False)
        return lobby

    def get_newroles(self, team):
        if team == 'soldier':
            queencheck = '✅' if 'queen' in self.newroles else '✖'
            ackermancheck = '✅' if 'ackerman' in self.newroles else '✖'
            mikecheck = '✅' if 'mike' in self.newroles else '✖'
            scoutcheck = '✅' if 'scout' in self.newroles else '✖'
            role_msg = queencheck + ' 👼**Queen**👼\n\n' + ackermancheck + ' 💂**Ackerman**💂\n\n' + \
                mikecheck + ' <:aotSmirk2:455223780957224961> **Mike Zacharias** <:aotSmirk2:455223780957224961>\n\n' + scoutcheck + ' 🏇**Scout**🏇'
            return role_msg
        elif team == 'warrior':
            warchiefcheck = '✅' if 'warchief' in self.newroles else '✖'
            falsekingcheck = '✅' if 'false king' in self.newroles else '✖'
            ymircheck = '✅' if 'ymir' in self.newroles else '✖'
            spycheck = '✅' if 'spy' in self.newroles else '✖'
            role_msg = warchiefcheck + ' 🦹‍♂️**Warchief**🦹‍♂️\n\n' + falsekingcheck + ' 🕴**False King**🕴\n\n' + ymircheck + ' 🤷‍♀️**Ymir**🤷‍♀️\n\n' + spycheck + ' 🕵️‍♀️**Spy**🕵️‍♀️'
            return role_msg
        else:
            ymirblessingcheck = '🔮' if self.ymir_blessing else '◼'
            pathscheck = '📢' if self.paths else '◼'
            role_msg = ymirblessingcheck + ' **Ymir\'s Blessing** ' + ymirblessingcheck + '\n\n' + pathscheck + ' **Paths** ' + pathscheck
            return role_msg

    def start(self, player):
        if self.status == 'waiting for players':
            if player == self.game_host:
                if len(self.players) >= 5:
                    self.message_box = []
                    self.status = 'assigning roles' # First input is selection of expedition members by commander

                    # Randomize roles
                    if self.randomroles:
                        self.randomize_roles()

                    # Add optional roles
                    all_player_roles = self.player_roles[len(self.players)].copy()
                    if 'queen' in self.newroles:
                        for i in range(len(all_player_roles)):
                            if all_player_roles[i] == 'soldier':
                                all_player_roles[i] = 'queen'
                                break
                    if 'ackerman' in self.newroles:
                        for i in range(len(all_player_roles)):
                            if all_player_roles[i] == 'soldier':
                                all_player_roles[i] = 'ackerman'
                                break
                    if 'mike' in self.newroles:
                        for i in range(len(all_player_roles)):
                            if all_player_roles[i] == 'soldier':
                                all_player_roles[i] = 'mike'
                                break
                    if 'scout' in self.newroles:
                        for i in range(len(all_player_roles)):
                            if all_player_roles[i] == 'soldier':
                                all_player_roles[i] = 'scout'
                                break
                    if 'warchief' in self.newroles:
                        for i in range(len(all_player_roles)):
                            if all_player_roles[i] == 'warrior':
                                all_player_roles[i] = 'warchief'
                                break
                    if 'ymir' in self.newroles:
                        for i in range(len(all_player_roles)):
                            if all_player_roles[i] == 'warrior':
                                all_player_roles[i] = 'ymir'
                                break
                    if 'false king' in self.newroles:
                        for i in range(len(all_player_roles)):
                            if all_player_roles[i] == 'warrior':
                                all_player_roles[i] = 'false king'
                                break
                    if 'spy' in self.newroles:
                        for i in range(len(all_player_roles)):
                            if all_player_roles[i] == 'warrior':
                                all_player_roles[i] = 'spy'
                                break
                    
                    self.players.sort(key = lambda x: random.randint(1,100)) # Randomizes player list
                    for i in range(len(self.players)): # Assigns roles to players
                        self.players[i][1] = all_player_roles[i]
                    self.players.sort(key = lambda x: random.randint(1,100)) # Randomizes player list again to get commander

                    if self.paths:
                        self.paths_holders = list(map(lambda x: x[0], self.players))
                        self.paths_holders.sort(key = lambda x: random.randint(1,100))

                    self.cur_expedition += 1
                    return 'Starting game with **' + str(len(self.players)) + '** players!'

                else:
                    return 'Need more players (at least 5 total) to start the game!'
            else:
                return 'Only the host, **' + self.game_host.mention + '**, can start the game!'

        elif self.status == 'waiting for game' or self.status.startswith('game ended'):
            return 'There is no open lobby at the moment!'
        else:
            return 'The game has already started!'

    def reset(self, player):
        if (player == self.game_host or player.id == 238808836075421697) or (self.game_host == None) or (self.status == 'waiting for players') or (self.status.startswith('game ended')):
            self.status = "waiting for game"

            self.game_host = None
            self.gametype = 'Ranked'
            self.players = []
            self.message_box = []
            self.newroles = []
            self.ymir_blessing = False
            self.blessed = []
            self.currently_blessed = None
            self.randomroles = False
            self.paths = False
            self.paths_holders = []

            self.cur_expedition = 0
            self.expedition_squad = []
            self.expedition_approval = []
            self.votes_flipped = False
            self.flipping_votes = False
            self.expedition_result = []

            self.expedition_history = []
            self.successful_expeditions = 0
            self.walls_destroyed = 0
            self.wall_secured = False

            return self.reset_msg

        else:
            return 'Only the host of the current game, **' + self.game_host.name + '**, can reset the game!'

    def get_role_msg(self, player_lst):
        soldier_msg = 'You are a 🛡**Soldier**🛡!\n\n\
As a member of the Survey Corps, you have been tasked with the responsibility of finding the secrets to the world \
by reaching the basement.\n\n\
Your task is to find out who the traitors within the military are, reach the basement before they destroy all 3 Walls, and help protect the identity \
of the Coordinate.'

        warrior_msg = 'You are a ⚔**Warrior**⚔!\n\n\
You are a Titan shifter who has infiltrated the Walls and are now working undercover to undermine the military\'s efforts. Working with your \
fellow Warriors, you seek to destroy the Walls and wipe out humanity. \n\n\
Your task is to conceal your identity from the Soldiers, and persuade them to allow you into their expeditions. At the same time, you need to try and identify the Coordinate within the ranks of the soldiers. \
If the Soldiers reach the basement before you destroy all the Walls, you can still achieve victory by kidnapping the Coordinate and bringing them to your hometown.\n\n\
Your fellow Warriors are:\n'
        for warrior in list(filter(lambda x:x[1] in self.warrior_roles, self.players)):
            if warrior == player_lst or warrior[1] == 'ymir':
                continue
            else:
                warrior_msg = warrior_msg + '**' + warrior[0].name + '**\n'

        coordinate_msg = 'You are the 🗺**Coordinate**🗺!\n\n\
You possess the special ability to sense Titan powers, which allows you to know who the Warriors are. However, due to the powers you possess, \
you are also their prime target, and so you must endeavour to remain hidden while aiding the Soldiers as much as you can. \n\n\
Your task is to help the Soldiers achieve victory using your knowledge of the Warriors\' identities, while keeping your identity as the Coordinate a secret from the Warriors.\n\n\
The Warriors are:\n'
        for warrior in list(filter(lambda x:x[1] in self.warrior_roles, self.players)):
            if warrior[1] == 'warchief':
                continue
            else:
                coordinate_msg = coordinate_msg + '**' + warrior[0].name + '**\n'

        queen_msg = 'You are the 👼**Queen**👼!\n\n\
As Queen of the Walls, you possess the ability to sense the Coordinate\'s powers. Use this knowledge wisely to help the Soldiers achieve victory.\n\n\
Your task is to protect the Coordinate by using your knowledge of their identity to deceive the Warriors into thinking you are the Coordinate.\n\n'
        if 'false king' not in self.newroles:
            queen_msg = queen_msg + 'The Coordinate is:\n**' + list(filter(lambda x:x[1] == 'coordinate', self.players))[0][0].name + '**'
        else:
            queen_msg = queen_msg + 'The Coordinate is one of the following two players. The other is the False King:\n'
            coord_list = ['coordinate', 'false king']
            coord_list.sort(key = lambda x: random.randint(1,100))
            for role in coord_list:
                queen_msg = queen_msg + '**' + list(filter(lambda x:x[1] == role, self.players))[0][0].name + '**\n'

        ackerman_msg = 'You are the 💂**Ackerman**💂!\n\n\
As one of humanity\'s strongest soldiers, you have been tasked with the responsibility of protecting the Walls.\n\n\
Your task is the same as the Soldiers: to find out who the traitors within the military are, reach the basement before they destroy all 3 Walls, and help protect the identity \
of the Coordinate.\n\n\
In addition, you also have the ability to **secure** the Walls on one expedition if you\'re in it. This prevents any Warriors on the expedition from destroying it. However, doing so also \
alerts any Warriors on the expedition of your identity as the Ackerman, making it easier for them to identity the Coordinate. Use this ability wisely.'

        mike_msg = 'You are <:aotSmirk2:455223780957224961>**Mike Zacharias** <:aotSmirk2:455223780957224961>!\n\n\
As humanity\'s 2nd strongest soldier, you have the ability to use your incredible sense of smell to sniff out Titans in expeditions.\n\n\
During the approval phase of each expedition, you will be told how many Titans the expedition team contains, if any. Both Warriors and the Coordinate are considered Titans.'

        scout_msg = 'You are the 🏇**Scout**🏇!\n\n\
As the guide of the expedition, you will fire a signal flare in any expedition that you are in, alerting all to your presence.\n\n\
Your task is to guide the Soldiers to pick the correct team for expeditions, while also hiding your identity from the Warriors to help protect the Coordinate.'

        warchief_msg = 'You are the 🦹‍♂️**Warchief**🦹‍♂️!\n\n\
As the leader of the Warriors, you have the unique ability of concealing your Titan powers from the Coordinate. Use this extra layer of anonymity to infiltrate deep into the Soldiers\' \
ranks and ensure the Warriors emerge victorious.\n\n\
Your task is the same as your fellow Warriors: Destroy the Walls by getting into expeditions, and try to identify the Coordinate.\n\n\
Your fellow Warriors are:\n'
        for warrior in list(filter(lambda x:x[1] in self.warrior_roles, self.players)):
            if warrior == player_lst or warrior[1] == 'ymir':
                continue
            else:
                warchief_msg = warchief_msg + '**' + warrior[0].name + '**\n'

        ymir_msg = 'You are 🤷‍♀️**Ymir**🤷‍♀️!\n\n\
While you are on the Warriors\' side, you do not know of their identities, nor do they know of yours. To achieve success, be sure to keep a close watch \
for potential comrades during expeditions.\n\n\
Your task is to identify your fellow Warriors, and work with them to destroy the Walls.'

        falseking_msg = 'You are the 🕴**False King**🕴!\n\n\
You appear as the Coordinate to the Queen, in addition to the real Coordinate. Use this ability to confuse her by acting as the Coordinate, to \
help turn the tide in the Warriors\' favor.\n\n\
Your task is the same as your fellow Warriors: Destroy the Walls by getting into expeditions, and try to identify the Coordinate.\n\n\
Your fellow Warriors are:\n'
        for warrior in list(filter(lambda x:x[1] in self.warrior_roles, self.players)):
            if warrior == player_lst or warrior[1] == 'ymir':
                continue
            else:
                falseking_msg = falseking_msg + '**' + warrior[0].name + '**\n'

        spy_msg = 'You are the 🕵️‍♀️**Spy**🕵️‍♀️!\n\n\
As the Warriors\' most skilled infiltrator, you have the ability to flip the votes during the approval phase of an expedition once.\n\n\
Your task is to help swing the expeditions in the Warriors\' favor at the right time, to achieve your goal of destroying the Walls.\n\n\
Your fellow Warriors are:\n'
        for warrior in list(filter(lambda x:x[1] in self.warrior_roles, self.players)):
            if warrior == player_lst or warrior[1] == 'ymir':
                continue
            else:
                spy_msg = spy_msg + '**' + warrior[0].name + '**\n'

        player_role = player_lst[1]
        role_msgs = {
            'soldier': soldier_msg, 
            'warrior': warrior_msg, 
            'coordinate': coordinate_msg,
            'queen': queen_msg, 
            'ackerman': ackerman_msg,
            'mike': mike_msg,
            'scout': scout_msg,
            'warchief': warchief_msg, 
            'ymir': ymir_msg, 
            'false king': falseking_msg,
            'spy': spy_msg
        }
        return role_msgs[player_role]

    def get_player_role(self, player = None):
        # Lists all roles currently in the game
        if player == None:
            # Game hasn't started yet
            if self.status == 'waiting for players' or self.status == 'waiting for game' or self.status.startswith('game ended'):
                soldier_roles = '\n'.join(['🛡**Soldier**🛡', '🗺**Coordinate**🗺', '👼**Queen**👼', '💂**Ackerman**💂', 
                '<:aotSmirk2:455223780957224961> **Mike Zacharias** <:aotSmirk2:455223780957224961>', '🏇**Scout**🏇'])
                warrior_roles = '\n'.join(['⚔**Warrior**⚔', '🦹‍♂️**Warchief**🦹‍♂️', '🕴**False King**🕴', '🤷‍♀️**Ymir**🤷‍♀️', '🕵️‍♀️**Spy**🕵️‍♀️'])
                
                list_of_roles = discord.Embed(title = 'List of roles available', description = 'Type ~help <role> for more information about a role.', colour=0x0013B4)
                list_of_roles.add_field(name = 'Soldier roles', value = soldier_roles, inline = True)
                list_of_roles.add_field(name = 'Warrior roles', value = warrior_roles, inline = True)
                return list_of_roles
            else:
                all_roles = [player[1] for player in self.players]

                def role_ordering(role_name):
                    role_map = {'soldier': 0, 'warrior': 1, 'coordinate': 2, 'queen': 3, 'ackerman': 4, 'mike': 5, 'scout': 6, 'warchief': 7, 'false king': 8, 'ymir': 9, 'spy': 10}
                    return role_map[role_name]

                all_roles.sort(key = role_ordering)
                display_roles = ''
                display_role_map = {'soldier': '🛡**Soldier**🛡', 'queen': '👼**Queen**👼', 'ackerman': '💂**Ackerman**💂', 
                'mike': '<:aotSmirk2:455223780957224961> **Mike Zacharias** <:aotSmirk2:455223780957224961>', 'scout': '🏇**Scout**🏇', 'coordinate': '🗺**Coordinate**🗺', 
                'warrior': '⚔**Warrior**⚔', 'warchief': '🦹‍♂️**Warchief**🦹‍♂️', 'false king': '🕴**False King**🕴', 'ymir': '🤷‍♀️**Ymir**🤷‍♀️', 'spy': '🕵️‍♀️**Spy**🕵️‍♀️'}
                for role in all_roles:
                    display_roles += display_role_map[role] + '\n'

                list_of_roles = discord.Embed(title = 'Current list of roles in game', description = display_roles, colour=0xF9FF41)
                return list_of_roles

        # DM player their role
        else:
            if player not in list(map(lambda x: x[0], self.players)):
                return 'You are not in a game!'
            elif self.status == 'waiting for players':
                return 'The game hasn\'t started yet!'
            else:
                player_role = list(filter(lambda x: player == x[0], self.players))[0][1]

                soldier_msg = 'You are a 🛡**Soldier**🛡!'

                warrior_msg = 'You are a ⚔**Warrior**⚔!\n\nYour fellow Warriors are:\n'
                for warrior in list(filter(lambda x:x[1] in self.warrior_roles, self.players)):
                    if warrior[0] == player or warrior[1] == 'ymir':
                        continue
                    else:
                        warrior_msg = warrior_msg + '**' + warrior[0].name + '**\n'

                warchief_msg = 'You are the 🦹‍♂️**Warchief**🦹‍♂️!\n\nYour fellow Warriors are:\n'
                for warrior in list(filter(lambda x:x[1] in self.warrior_roles, self.players)):
                    if warrior[0] == player or warrior[1] == 'ymir':
                        continue
                    else:
                        warchief_msg = warchief_msg + '**' + warrior[0].name + '**\n'

                queen_msg = 'You are the 👼**Queen**👼!\n\n'
                if 'false king' not in self.newroles:
                    queen_msg = queen_msg + 'The Coordinate is:\n**' + list(filter(lambda x:x[1] == 'coordinate', self.players))[0][0].name + '**'
                else:
                    queen_msg = queen_msg + 'The Coordinate is one of the following two players. The other is the False King:\n'
                    coord_list = ['coordinate', 'false king']
                    coord_list.sort(key = lambda x: random.randint(1,100))
                    for role in coord_list:
                        queen_msg = queen_msg + '**' + list(filter(lambda x:x[1] == role, self.players))[0][0].name + '**\n'

                ackerman_msg = 'You are the 💂**Ackerman**💂!'

                mike_msg = 'You are <:aotSmirk2:455223780957224961>**Mike Zacharias** <:aotSmirk2:455223780957224961>!'

                scout_msg = 'You are the 🏇**Scout**🏇!'
                
                coordinate_msg = 'You are the 🗺**Coordinate**🗺!\n\nThe Warriors are:\n'
                for warrior in list(filter(lambda x:x[1] in self.warrior_roles, self.players)):
                    if warrior[1] == 'warchief':
                        continue
                    else:
                        coordinate_msg = coordinate_msg + '**' + warrior[0].name + '**\n'

                ymir_msg = 'You are 🤷‍♀️**Ymir**🤷‍♀️!'

                falseking_msg = 'You are the 🕴**False King**🕴!\n\nYour fellow Warriors are:\n'
                for warrior in list(filter(lambda x:x[1] in self.warrior_roles, self.players)):
                    if warrior[0] == player or warrior[1] == 'ymir':
                        continue
                    else:
                        falseking_msg = falseking_msg + '**' + warrior[0].name + '**\n'

                spy_msg = 'You are the 🕵️‍♀️**Spy**🕵️‍♀️!\n\nYour fellow Warriors are:\n'
                for warrior in list(filter(lambda x:x[1] in self.warrior_roles, self.players)):
                    if warrior[0] == player or warrior[1] == 'ymir':
                        continue
                    else:
                        spy_msg = spy_msg + '**' + warrior[0].name + '**\n'

                role_msgs = {
                    'soldier': soldier_msg, 
                    'warrior': warrior_msg, 
                    'coordinate': coordinate_msg,
                    'queen': queen_msg, 
                    'ackerman': ackerman_msg, 
                    'mike': mike_msg, 
                    'scout': scout_msg,
                    'warchief': warchief_msg, 
                    'ymir': ymir_msg, 
                    'false king': falseking_msg,
                    'spy': spy_msg
                }
                return role_msgs[player_role]

    def pick(self, player):
        if player in list(map(lambda x: x[0], self.players)):
            if player in self.expedition_squad:
                return '**' + player.name + '** is already part of the expedition team!'
            else:
                self.expedition_squad.append(player)
                return '**' + player.name + '** has joined the expedition team!'
        else:
            return '**' + player.name + '** is not in the game!'

    def get_ymir_blessing(self):
        if self.currently_blessed == None:
            self.currently_blessed = random.choice(list(filter(lambda x: x not in self.blessed, list(map(lambda x: x[0], self.players)))))
        return self.currently_blessed.mention + ' has received Ymir\'s blessing!'
    
    def get_player_allegiance(self, player):
        if player in list(map(lambda x:x[0], list(filter(lambda x:x[1] in self.warrior_roles, self.players)))):
            return '⚔**Warriors**⚔!'
        else:
            return '🛡**Soldiers**🛡!'

    def get_expedition_size(self):
        expedition_size = self.expedition_sizes[len(self.players)][self.cur_expedition-1]
        return expedition_size

    def next_commander(self):
        prev_commander = self.players.pop(0)
        self.players.append(prev_commander)
        self.message_box = []

    def selection_players(self):
        players_in_expedition = ''
        for person in self.expedition_squad:
            players_in_expedition = players_in_expedition + '**' + person.name + '**\n'
        selection_state = discord.Embed(title = '**' + str(len(self.expedition_squad)) + '**/**' + str(self.get_expedition_size()) + 
        '** players in expedition team', description = players_in_expedition, colour=0xF9FF41)
        return selection_state

    def begin_approval_phase(self):
        self.status = 'expedition approval'
        self.message_box = []
        return '--**Expedition #' + str(self.cur_expedition) + ': Approval phase**--'

    def approval_players(self):
        approval_status = ''
        for player in list(map(lambda x:x[0], self.players)):
            if player in list(map(lambda x:x[0], self.expedition_approval)):
                approval_status = approval_status + '**' + player.name + '**\n'
            else:
                approval_status = approval_status + player.name + '\n'
        expedition_approval_status = discord.Embed(title = 'Expedition approval', description = approval_status, colour=0xF9FF41)
        return expedition_approval_status
    
    def approval_add_remainder(self):
        for player in list(map(lambda x:x[0], self.players)):
            if player not in list(map(lambda x:x[0], self.expedition_approval)):
                self.expedition_approval.append([player, 'yes'])

    def flip_votes(self):
        for player in self.expedition_approval:
            if player[1] == 'yes':
                player[1] = 'no'
            elif player[1] == 'no':
                player[1] = 'yes'

        self.flipping_votes = False

    def get_approval_result(self):
        approval_result = ''
        for player in self.expedition_approval:
            if player[1] == 'yes':
                approval_result = approval_result + '**' + player[0].name + '** ✅\n'
            elif player[1] == 'no':
                approval_result = approval_result + '**' + player[0].name + '** ❌\n'
        expedition_result = discord.Embed(title = 'Expedition approval result', description = approval_result, colour=0xF9FF41)
        return expedition_result

    def decision_players(self):
        decision_status = str(len(list(filter(lambda x: x in list(map(lambda x:x[0], self.expedition_result)), self.expedition_squad)))) + \
            '/' + str(len(self.expedition_squad)) + ' players'

        expedition_decision_status = discord.Embed(title = 'Expedition decision', description = decision_status, colour=0xF9FF41)
        return expedition_decision_status

    def scout_in_expedition(self):
        players_in_expedition_squad = list(filter(lambda x: x[0] in self.expedition_squad, self.players))
        return 'scout' in list(map(lambda x: x[1], players_in_expedition_squad))

    def get_expedition_result(self):
        self.status = ''
        decision_result = ''
        self.expedition_result.sort(key = lambda x: random.randint(1,100))
        if 'secure' in list(map(lambda x:x[1], self.expedition_result)):
            for player in self.expedition_result:
                decision_result = decision_result + '✅\n'
        else:
            for player in self.expedition_result:
                if player[1] == 'yes':
                    decision_result = decision_result + '✅\n'
                elif player[1] == 'no':
                    decision_result = decision_result + '❌\n'
        
        expedition_result = discord.Embed(title = 'Expedition Result', description = decision_result, colour=0xF9FF41)
        return expedition_result

    def test_expedition_success(self):
        num_fails = len(list(filter(lambda x: x[1] == 'no', self.expedition_result)))
        if 'secure' in list(map(lambda x:x[1], self.expedition_result)):
            num_fails = 0

        if len(self.players) >= 7 and self.cur_expedition == 4:
            if num_fails >= 2:
                return False
            else:
                return True
        else:
            if num_fails >= 1:
                return False
            else:
                return True

    def expedition_success_update(self):
        self.successful_expeditions += 1
        self.cur_expedition += 1
        self.expedition_history.append('🛡')
        self.expedition_squad = []
        self.expedition_result = []
        return '✅The expedition was a success!✅'

    def expedition_failure_update(self):
        self.walls_destroyed += 1
        self.cur_expedition += 1
        self.expedition_history.append('⚔')
        self.expedition_squad = []
        self.expedition_result = []

        result_msg = 'The expedition was a failure!\n\n'
        if self.walls_destroyed == 1:
            result_msg = result_msg + '💥Wall **Maria** has been destroyed!💥'
        elif self.walls_destroyed == 2:
            result_msg = result_msg + '💥Wall **Rose** has been destroyed!💥'
        elif self.walls_destroyed == 3:
            result_msg = result_msg + '💥Wall **Sina** has been destroyed!💥'
        return result_msg

    def reach_basement(self):
        self.status = 'choose coordinate'
        basement_msg = 'The Soldiers have reached the Basement! The Warriors have one last chance to seize victory by capturing the Coordinate. Any Warrior may type **~kidnap <@name>** \
to select who they believe is the Coordinate.\n\nThe Warriors are:\n'
        for person in list(filter(lambda x:x[1] in self.warrior_roles, self.players)):
            basement_msg += '**' + person[0].name + '**\n'
        return basement_msg

    def addendum(self, number):
        if self.cur_expedition == number:
            return '🏇'
        elif self.cur_expedition < number:
            return ''
        elif self.cur_expedition > number:
            return self.expedition_history[number-1]

    def add2fails(self):
        if len(self.players) >= 7:
            return '(2 Warriors to fail)'
        else:
            return ''

    def get_status(self):
        if self.status == 'waiting for game' or self.status == 'waiting for players':
            return None
        else:
            status = discord.Embed(title='Current Game Information', description= 'Number of players: **' + str(len(self.players)) + '**\n**' + 
            str(len(list(filter(lambda x:x[1] not in self.warrior_roles, self.players)))) + 
            '** Soldiers vs **' + str(len(list(filter(lambda x:x[1] in self.warrior_roles, self.players)))) + '** Warriors', colour=0xF9FF41)

            if self.successful_expeditions == 0:
                basement_progress = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ \u2001\u2001\u20011⃣\n\u2001\u2001\u2001\u2001\u20012⃣\n\u2001\u2001\u2001\u2001\u20013⃣🏁'
            elif self.successful_expeditions == 1:
                basement_progress = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ \u2001\u2001\u2001✅\n\u2001\u2001\u2001\u2001\u20012⃣\n\u2001\u2001\u2001\u2001\u20013⃣🏁'
            elif self.successful_expeditions == 2:
                basement_progress = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ \u2001\u2001\u2001✅\n\u2001\u2001\u2001\u2001\u2001✅\n\u2001\u2001\u2001\u2001\u20013⃣🏁'
            elif self.successful_expeditions == 3:
                basement_progress = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ \u2001\u2001\u2001✅\n\u2001\u2001\u2001\u2001\u2001✅\n\u2001\u2001\u2001\u2001\u2001✅🏁'

            if self.walls_destroyed == 0:
                wall_status = '￶￵ ￶￵  ￶￵ ￶￵ ￶￵ ￶￵ \u200A\u200A\u2000🏠🇲🏠\n\u2001\u2000\u2000🏦🇷🏦\n\u2001\u2000\u2000🏰🇸🏰'
            elif self.walls_destroyed == 1:
                wall_status = '￶￵ ￶￵  ￶￵ ￶￵ ￶￵ ￶￵ \u200A\u200A\u2000🏚💥🏚\n\u2001\u2000\u2000🏦🇷🏦\n\u2001\u2000\u2000🏰🇸🏰'
            elif self.walls_destroyed == 2:
                wall_status = '￶￵ ￶￵  ￶￵ ￶￵ ￶￵ ￶￵ \u200A\u200A\u2000🏚💥🏚\n\u2001\u2000\u2000🏚💥🏚\n\u2001\u2000\u2000🏰🇸🏰'
            elif self.walls_destroyed >= 3:
                wall_status = '￶￵ ￶￵  ￶￵ ￶￵ ￶￵ ￶￵ \u200A\u200A\u2000🏚💥🏚\n\u2001\u2000\u2000🏚💥🏚\n\u2001\u2000\u2000🏚💥🏚'

            expedition_info = 'Expedition 1: **' + str(self.expedition_sizes[len(self.players)][0]) + '** members ' + self.addendum(1) + '\n\n\
Expedition 2: **' + str(self.expedition_sizes[len(self.players)][1]) + '** members ' + self.addendum(2) + '\n\n\
Expedition 3: **' + str(self.expedition_sizes[len(self.players)][2]) + '** members ' + self.addendum(3) + '\n\n\
Expedition 4: **' + str(self.expedition_sizes[len(self.players)][3]) + '** members ' + self.add2fails() + self.addendum(4) + '\n\n\
Expedition 5: **' + str(self.expedition_sizes[len(self.players)][4]) + '** members ' + self.addendum(5)

            status.add_field(name = 'Progress towards Basement', value = basement_progress, inline = True)
            status.add_field(name = 'Status of the Walls', value = wall_status, inline = True)
            status.add_field(name = 'Expedition Info', value = expedition_info, inline = False)

            return status
    
    def get_players(self):
        if self.status != 'waiting for game' and self.status != 'waiting for players':
            all_players = ''
            for player in self.players:
                if player == self.players[0]:
                    all_players = all_players + '**' + player[0].mention + '** 👑\n'
                else:
                    all_players = all_players + '**' + player[0].mention + '**' + '\n'
            players = discord.Embed(title = 'List of players (in queue to become Commander)', description = all_players, colour=0xF9FF41)
            return players
        elif self.status == 'waiting for players':
            return self.display_lobby()
        else:
            return None

    def get_summary(self):
        summary = discord.Embed(title = 'Game Summary', description = 'Number of players: **' + str(len(self.players)) + '**\n**' + 
str(len(list(filter(lambda x:x[1] not in self.warrior_roles, self.players)))) + 
'** Soldiers vs **' + str(len(list(filter(lambda x:x[1] in self.warrior_roles, self.players)))) + '** Warriors', colour=0xF9FF41)

        all_soldiers = ''
        all_warriors = ''
        for person in self.players:
            if person[1] == 'soldier':
                all_soldiers += person[0].name + '\n'
            elif person[1] == 'warrior':
                all_warriors += person[0].name + '\n'

        if all_soldiers:
            summary.add_field(name = '🛡Soldiers🛡', value = all_soldiers)
        if all_warriors:
            summary.add_field(name = '⚔Warriors⚔', value = all_warriors)
        summary.add_field(name = '🗺Coordinate🗺', value = list(filter(lambda x:x[1] == 'coordinate', self.players))[0][0].name)
        if 'queen' in self.newroles:
            summary.add_field(name = '👼Queen👼', value = list(filter(lambda x:x[1] == 'queen', self.players))[0][0].name)
        if 'ackerman' in self.newroles:
            summary.add_field(name = '💂Ackerman💂', value = list(filter(lambda x:x[1] == 'ackerman', self.players))[0][0].name)
        if 'mike' in self.newroles:
            summary.add_field(name = '<:aotSmirk2:455223780957224961> Mike Zacharias <:aotSmirk2:455223780957224961>', value = list(filter(lambda x:x[1] == 'mike', self.players))[0][0].name)    
        if 'scout' in self.newroles:
            summary.add_field(name = '🏇Scout🏇', value = list(filter(lambda x:x[1] == 'scout', self.players))[0][0].name)
        if 'warchief' in self.newroles:
            summary.add_field(name = '🦹‍♂️Warchief🦹‍♂️', value = list(filter(lambda x:x[1] == 'warchief', self.players))[0][0].name)
        if 'false king' in self.newroles:
            summary.add_field(name = '🕴False King🕴', value = list(filter(lambda x:x[1] == 'false king', self.players))[0][0].name)
        if 'ymir' in self.newroles:
            summary.add_field(name = '🤷‍♀️Ymir🤷‍♀️', value = list(filter(lambda x:x[1] == 'ymir', self.players))[0][0].name)
        if 'spy' in self.newroles:
            summary.add_field(name = '🕵️‍♀️Spy🕵️‍♀️', value = list(filter(lambda x:x[1] == 'spy', self.players))[0][0].name)

        return summary

    def calculate_advantage(self, write=True, emblems=True):
        num_players = max(len(self.players), 5)
        expedition_advantage = self.expedition_advantages[num_players]
        roles_advantage = 0
        for role in self.newroles:
            roles_advantage += self.roles_advantages[role][num_players]
        advantage_score = roles_advantage + expedition_advantage
        if self.ymir_blessing:
            advantage_score += self.blessing_advantages[num_players]

        if write == False:
            return advantage_score
        if advantage_score == 0:
            return '**None**'
        elif advantage_score > 0:
            if emblems:
                return '🛡 **Soldiers (+%d)** 🛡' % advantage_score
            else:
                return '**Soldiers (+%d)**' % advantage_score
        else:
            if emblems:
                return '⚔ **Warriors (+%d)** ⚔' % -advantage_score
            else:
                return '**Warriors (+%d)**' % -advantage_score

    def update_rating(self):
        # Updates rating of all players after game ends, and returns it in an embed. Updates game records as well.
        wb = load_workbook("WarriorsvsSoldiers/database.xlsx")
        player_data = wb['Player records']
        game_data = wb['Game records']

        # Update game records
        num_players = len(self.players)
        player_dict = {5:'B', 6:'C', 7:'D', 8:'E', 9:'F', 10:'G'}
        if self.status == 'game ended soldiers':
            cell = player_dict[num_players] + '2'
            cell2 = 'H2'
            cell3 = player_dict[num_players] + '10'
            cell4 = 'H10'
            game_data[cell] = game_data[cell].value + 1
            game_data[cell2] = game_data[cell2].value + 1
            game_data[cell3] = game_data[cell3].value + 1
            game_data[cell4] = game_data[cell4].value + 1
            
        elif self.status == 'game ended warriors wall':
            cell = player_dict[num_players] + '4'
            cell2 = 'H4'
            cell3 = player_dict[num_players] + '8'
            cell4 = player_dict[num_players] + '10'
            cell5 = 'H8'
            cell6 = 'H10'
            game_data[cell] = game_data[cell].value + 1
            game_data[cell2] = game_data[cell2].value + 1
            game_data[cell3] = game_data[cell3].value + 1
            game_data[cell4] = game_data[cell4].value + 1
            game_data[cell5] = game_data[cell5].value + 1
            game_data[cell6] = game_data[cell6].value + 1

        elif self.status == 'game ended warriors coord':
            cell = player_dict[num_players] + '6'
            cell2 = 'H6'
            cell3 = player_dict[num_players] + '8'
            cell4 = player_dict[num_players] + '10'
            cell5 = 'H8'
            cell6 = 'H10'
            game_data[cell] = game_data[cell].value + 1
            game_data[cell2] = game_data[cell2].value + 1
            game_data[cell3] = game_data[cell3].value + 1
            game_data[cell4] = game_data[cell4].value + 1
            game_data[cell5] = game_data[cell5].value + 1
            game_data[cell6] = game_data[cell6].value + 1

        # Update player ratings
        ratings = []
        # Obtain original ratings
        for player in self.players:
            player_id = str(player[0].id)
            inside = False
            for row in player_data:
                if row[0].value == player_id:
                    ratings.append([player_id, row[1].value, player[1]]) # [id, rating, role]
                    inside = True
                    break
            if inside == False:
                player_data.append([player_id, 1500] + 30 * [0])
                ratings.append([player_id, 1500, player[1]])
        
        avg_s = sum(map(lambda y: y[1], (filter(lambda x: x[2] not in self.warrior_roles, ratings)))) / len(list(filter(lambda x: x[2] not in self.warrior_roles, ratings)))
        avg_w = sum(map(lambda y: y[1], (filter(lambda x: x[2] in self.warrior_roles, ratings)))) / len(list(filter(lambda x: x[2] in self.warrior_roles, ratings)))
        sr_modifier = self.calculate_advantage(write=False)

        # Calculate rating transfer amount
        K = 64
        exp = 1 / (1 + 10 ** ((avg_w - avg_s - sr_modifier)/800))
        if self.status == 'game ended soldiers':
            rating_transfer = K * (1 - exp)
        else:
            rating_transfer = K * exp

        # Apply rating transfers and update player records
        rating_changes = [] # [id, old rating, new rating, role]
        for player in ratings:
            if self.status == 'game ended soldiers':
                if player[2] not in self.warrior_roles:
                    new_rating = player[1] + rating_transfer
                    win = True
                else:
                    new_rating = player[1] - rating_transfer
                    win = False
            else:
                if player[2] not in self.warrior_roles:
                    new_rating = player[1] - rating_transfer
                    win = False
                else:
                    new_rating = player[1] + rating_transfer
                    win = True

            rating_changes.append([player[0], player[1], new_rating, player[2]])
            
            i = 0
            for row in player_data:
                i += 1
                if row[0].value == player[0]:
                    # Update with new rating
                    player_data['B'+ str(i)] = new_rating

                    # Update win/played stats
                    if player[2] == 'soldier':
                        player_data['D'+ str(i)] = player_data['D'+ str(i)].value + 1
                        player_data['F'+ str(i)] = player_data['F'+ str(i)].value + 1
                        player_data['J'+ str(i)] = player_data['J'+ str(i)].value + 1
                        if win == True:
                            player_data['C'+ str(i)] = player_data['C'+ str(i)].value + 1
                            player_data['E'+ str(i)] = player_data['E'+ str(i)].value + 1
                            player_data['I'+ str(i)] = player_data['I'+ str(i)].value + 1
                    elif player[2] == 'queen':
                        player_data['D'+ str(i)] = player_data['D'+ str(i)].value + 1
                        player_data['F'+ str(i)] = player_data['F'+ str(i)].value + 1
                        player_data['P'+ str(i)] = player_data['P'+ str(i)].value + 1
                        if win == True:
                            player_data['C'+ str(i)] = player_data['C'+ str(i)].value + 1
                            player_data['E'+ str(i)] = player_data['E'+ str(i)].value + 1
                            player_data['O'+ str(i)] = player_data['O'+ str(i)].value + 1
                    elif player[2] == 'ackerman':
                        player_data['D'+ str(i)] = player_data['D'+ str(i)].value + 1
                        player_data['F'+ str(i)] = player_data['F'+ str(i)].value + 1
                        player_data['Z'+ str(i)] = player_data['Z'+ str(i)].value + 1
                        if win == True:
                            player_data['C'+ str(i)] = player_data['C'+ str(i)].value + 1
                            player_data['E'+ str(i)] = player_data['E'+ str(i)].value + 1
                            player_data['Y'+ str(i)] = player_data['Y'+ str(i)].value + 1
                    elif player[2] == 'mike':
                        player_data['D'+ str(i)] = player_data['D'+ str(i)].value + 1
                        player_data['F'+ str(i)] = player_data['F'+ str(i)].value + 1
                        player_data['AB'+ str(i)] = player_data['AB'+ str(i)].value + 1
                        if win == True:
                            player_data['C'+ str(i)] = player_data['C'+ str(i)].value + 1
                            player_data['E'+ str(i)] = player_data['E'+ str(i)].value + 1
                            player_data['AA'+ str(i)] = player_data['AA'+ str(i)].value + 1
                    elif player[2] == 'scout':
                        player_data['D'+ str(i)] = player_data['D'+ str(i)].value + 1
                        player_data['F'+ str(i)] = player_data['F'+ str(i)].value + 1
                        player_data['AD'+ str(i)] = player_data['AD'+ str(i)].value + 1
                        if win == True:
                            player_data['C'+ str(i)] = player_data['C'+ str(i)].value + 1
                            player_data['E'+ str(i)] = player_data['E'+ str(i)].value + 1
                            player_data['AC'+ str(i)] = player_data['AC'+ str(i)].value + 1
                    elif player[2] == 'coordinate':
                        player_data['D'+ str(i)] = player_data['D'+ str(i)].value + 1
                        player_data['F'+ str(i)] = player_data['F'+ str(i)].value + 1
                        player_data['N'+ str(i)] = player_data['N'+ str(i)].value + 1
                        if win == True:
                            player_data['C'+ str(i)] = player_data['C'+ str(i)].value + 1
                            player_data['E'+ str(i)] = player_data['E'+ str(i)].value + 1
                            player_data['M'+ str(i)] = player_data['M'+ str(i)].value + 1
                    elif player[2] == 'warrior':
                        player_data['D'+ str(i)] = player_data['D'+ str(i)].value + 1
                        player_data['H'+ str(i)] = player_data['H'+ str(i)].value + 1
                        player_data['L'+ str(i)] = player_data['L'+ str(i)].value + 1
                        if win == True:
                            player_data['C'+ str(i)] = player_data['C'+ str(i)].value + 1
                            player_data['G'+ str(i)] = player_data['G'+ str(i)].value + 1
                            player_data['K'+ str(i)] = player_data['K'+ str(i)].value + 1
                    elif player[2] == 'warchief':
                        player_data['D'+ str(i)] = player_data['D'+ str(i)].value + 1
                        player_data['H'+ str(i)] = player_data['H'+ str(i)].value + 1
                        player_data['R'+ str(i)] = player_data['R'+ str(i)].value + 1
                        if win == True:
                            player_data['C'+ str(i)] = player_data['C'+ str(i)].value + 1
                            player_data['G'+ str(i)] = player_data['G'+ str(i)].value + 1
                            player_data['Q'+ str(i)] = player_data['Q'+ str(i)].value + 1
                    elif player[2] == 'ymir':
                        player_data['D'+ str(i)] = player_data['D'+ str(i)].value + 1
                        player_data['H'+ str(i)] = player_data['H'+ str(i)].value + 1
                        player_data['V'+ str(i)] = player_data['V'+ str(i)].value + 1
                        if win == True:
                            player_data['C'+ str(i)] = player_data['C'+ str(i)].value + 1
                            player_data['G'+ str(i)] = player_data['G'+ str(i)].value + 1
                            player_data['U'+ str(i)] = player_data['U'+ str(i)].value + 1
                    elif player[2] == 'false king':
                        player_data['D'+ str(i)] = player_data['D'+ str(i)].value + 1
                        player_data['H'+ str(i)] = player_data['H'+ str(i)].value + 1
                        player_data['X'+ str(i)] = player_data['X'+ str(i)].value + 1
                        if win == True:
                            player_data['C'+ str(i)] = player_data['C'+ str(i)].value + 1
                            player_data['G'+ str(i)] = player_data['G'+ str(i)].value + 1
                            player_data['W'+ str(i)] = player_data['W'+ str(i)].value + 1
                    elif player[2] == 'spy':
                        player_data['D'+ str(i)] = player_data['D'+ str(i)].value + 1
                        player_data['H'+ str(i)] = player_data['H'+ str(i)].value + 1
                        player_data['AF'+ str(i)] = player_data['AF'+ str(i)].value + 1
                        if win == True:
                            player_data['C'+ str(i)] = player_data['C'+ str(i)].value + 1
                            player_data['G'+ str(i)] = player_data['G'+ str(i)].value + 1
                            player_data['AE'+ str(i)] = player_data['AE'+ str(i)].value + 1
                    
                    break

        wb.save("WarriorsvsSoldiers/database.xlsx")

        # Embed skill rating changes
        new_ratings = discord.Embed(title = 'Skill Rating (SR) Update', 
                                    description = 'Advantage: ' + self.calculate_advantage(emblems=False) + '\nSoldiers: **' + str(int(round(avg_s, 0))) + '** | Warriors: **' + str(int(round(avg_w, 0))) + '**', 
                                    colour=0xF9FF41)

        soldier_ratings = ''
        warrior_ratings = ''
        for player in self.players:
            if player[1] not in self.warrior_roles:
                player_info = list(filter(lambda x:x[0] == str(player[0].id), rating_changes))[0]
                if self.status == 'game ended soldiers':
                    sign = '+'
                else:
                    sign = '-'
                soldier_ratings += player[0].name + ': **' + str(int(round(player_info[1], 0))) + '** -> **' + str(int(round(player_info[2], 0))) + '** (' + sign + str(int(round(rating_transfer, 0))) + ')\n'

            else:
                player_info = list(filter(lambda x:x[0] == str(player[0].id), rating_changes))[0]
                if self.status == 'game ended soldiers':
                    sign = '-'
                else:
                    sign = '+'
                warrior_ratings += player[0].name + ': **' + str(int(round(player_info[1], 0))) + '** -> **' + str(int(round(player_info[2], 0))) + '** (' + sign + str(int(round(rating_transfer, 0))) + ')\n'

        new_ratings.add_field(name = '🛡Soldiers🛡', value = soldier_ratings, inline = False)
        new_ratings.add_field(name = '⚔Warriors⚔', value = warrior_ratings, inline = False)
        return new_ratings

    def update_achievements(self):
        achievements_msgs = []
        wb = load_workbook("WarriorsvsSoldiers/database.xlsx")
        player_data = wb['Player records']

        for cur_player in self.players:
            i = 0
            for row in player_data:
                i += 1
                # Player exists in the records (definitely, after updating rating)
                if row[0].value == str(cur_player[0].id):
                    if self.status == 'game ended soldiers':
                        if cur_player[1] not in self.warrior_roles:
                            win = True
                        else:
                            win = False
                    else:
                        if cur_player[1] not in self.warrior_roles:
                            win = False
                        else:
                            win = True

                    # Update consecutive wins and add SR, achievement msg if needed
                    if win == True:
                        player_data['S'+ str(i)] = player_data['S'+ str(i)].value + 1
                        if player_data['S'+ str(i)].value > player_data['T'+ str(i)].value:
                            # New win streak
                            player_data['T'+ str(i)] = player_data['S'+ str(i)].value

                            if player_data['T' + str(i)].value in self.consecutive_wins_achievements:
                                # Achievement msg
                                consecutive_wins_msg = self.badge_emojis[self.consecutive_wins_achievements[player_data['T' + str(i)].value]] + ' | Congratulations ' + cur_player[0].mention + \
                                '! You have obtained a **' + self.consecutive_wins_achievements[player_data['T' + str(i)].value] + '** badge for winning **' + \
                                str(player_data['T' + str(i)].value) + '** games in a row! **(+' + str(self.achievement_rewards[self.consecutive_wins_achievements[player_data['T' + str(i)].value]]) + ' SR)**'

                                achievements_msgs.append(consecutive_wins_msg)

                                # Add SR
                                player_data['B'+ str(i)] = player_data['B'+ str(i)].value + self.achievement_rewards[self.consecutive_wins_achievements[player_data['T' + str(i)].value]]

                    else:
                        player_data['S'+ str(i)] = 0

                    # Add SR, achievement msgs for other achievements if needed
                    if player_data['D'+ str(i)].value in self.number_games_achievements:
                        # Achievement msg
                        games_played_msg = self.badge_emojis[self.number_games_achievements[player_data['D' + str(i)].value]] + ' | Congratulations ' + cur_player[0].mention + \
                        '! You have obtained a **' + self.number_games_achievements[player_data['D' + str(i)].value] + '** badge for having played **' + \
                        str(player_data['D' + str(i)].value) + '** games! **(+' + str(self.achievement_rewards[self.number_games_achievements[player_data['D' + str(i)].value]]) + ' SR)**'

                        achievements_msgs.append(games_played_msg)

                        # Add SR
                        player_data['B'+ str(i)] = player_data['B'+ str(i)].value + self.achievement_rewards[self.number_games_achievements[player_data['D' + str(i)].value]]

                    if win == True:
                        if cur_player[1] == 'soldier':
                            if player_data['I' + str(i)].value in self.soldier_wins_achievements:
                                # Achievement msg
                                soldiers_win_msg = self.badge_emojis[self.soldier_wins_achievements[player_data['I' + str(i)].value]] + ' | Congratulations ' + cur_player[0].mention + \
                                '! You have obtained a **' + self.soldier_wins_achievements[player_data['I' + str(i)].value] + '** badge for winning **' + \
                                str(player_data['I' + str(i)].value) + '** games as a 🛡**Soldier**🛡! **(+' + str(self.achievement_rewards[self.soldier_wins_achievements[player_data['I' + str(i)].value]]) + ' SR)**'

                                achievements_msgs.append(soldiers_win_msg)

                                # Add SR
                                player_data['B'+ str(i)] = player_data['B'+ str(i)].value + self.achievement_rewards[self.soldier_wins_achievements[player_data['I' + str(i)].value]]

                        elif cur_player[1] == 'warrior':
                            if player_data['K' + str(i)].value in self.warrior_wins_achievements:
                                # Achievement msg
                                warriors_win_msg = self.badge_emojis[self.warrior_wins_achievements[player_data['K' + str(i)].value]] + ' | Congratulations ' + cur_player[0].mention + \
                                '! You have obtained a **' + self.warrior_wins_achievements[player_data['K' + str(i)].value] + '** badge for winning **' + \
                                str(player_data['K' + str(i)].value) + '** games as a ⚔**Warrior**⚔! **(+' + str(self.achievement_rewards[self.warrior_wins_achievements[player_data['K' + str(i)].value]]) + ' SR)**'

                                achievements_msgs.append(warriors_win_msg)

                                # Add SR
                                player_data['B'+ str(i)] = player_data['B'+ str(i)].value + self.achievement_rewards[self.warrior_wins_achievements[player_data['K' + str(i)].value]]

                        elif cur_player[1] == 'coordinate':
                            if player_data['M' + str(i)].value in self.coordinate_wins_achievements:
                                # Achievement msg
                                coordinate_win_msg = self.badge_emojis[self.coordinate_wins_achievements[player_data['M' + str(i)].value]] + ' | Congratulations ' + cur_player[0].mention + \
                                '! You have obtained a **' + self.coordinate_wins_achievements[player_data['M' + str(i)].value] + '** badge for winning **' + \
                                str(player_data['M' + str(i)].value) + '** games as the 🗺**Coordinate**🗺! **(+' + str(self.achievement_rewards[self.coordinate_wins_achievements[player_data['M' + str(i)].value]]) + ' SR)**'

                                achievements_msgs.append(coordinate_win_msg)

                                # Add SR
                                player_data['B'+ str(i)] = player_data['B'+ str(i)].value + self.achievement_rewards[self.coordinate_wins_achievements[player_data['M' + str(i)].value]]

                        elif cur_player[1] == 'queen' or cur_player[1] == 'ackerman' or cur_player[1] == 'mike' or cur_player[1] == 'scout':
                            total_wins = player_data['O' + str(i)].value + player_data['Y' + str(i)].value + player_data['AA' + str(i)].value + player_data['AC' + str(i)].value
                            if total_wins in self.queen_wins_achievements:
                                # Achievement msg
                                queen_win_msg = self.badge_emojis[self.queen_wins_achievements[total_wins]] + ' | Congratulations ' + cur_player[0].mention + \
                                '! You have obtained a **' + self.queen_wins_achievements[total_wins] + '** badge for winning **' + \
                                str(total_wins) + '** games with an optional Soldier role! **(+' + \
                                str(self.achievement_rewards[self.queen_wins_achievements[total_wins]]) + ' SR)**'

                                achievements_msgs.append(queen_win_msg)

                                # Add SR
                                player_data['B'+ str(i)] = player_data['B'+ str(i)].value + self.achievement_rewards[self.queen_wins_achievements[total_wins]]

                        elif cur_player[1] == 'warchief' or cur_player[1] == 'false king' or cur_player[1] == 'ymir' or cur_player[1] == 'spy':
                            total_wins = player_data['Q' + str(i)].value + player_data['W' + str(i)].value + player_data['U' + str(i)].value + player_data['AE' + str(i)].value
                            if total_wins in self.warchief_wins_achievements:
                                # Achievement msg
                                warchief_win_msg = self.badge_emojis[self.warchief_wins_achievements[total_wins]] + ' | Congratulations ' + cur_player[0].mention + \
                                '! You have obtained a **' + self.warchief_wins_achievements[total_wins] + '** badge for winning **' + \
                                str(total_wins) + '** games with an optional Warrior role! **(+' + str(self.achievement_rewards[self.warchief_wins_achievements[total_wins]]) + ' SR)**'

                                achievements_msgs.append(warchief_win_msg)

                                # Add SR
                                player_data['B'+ str(i)] = player_data['B'+ str(i)].value + self.achievement_rewards[self.warchief_wins_achievements[total_wins]]

                    wb.save("WarriorsvsSoldiers/database.xlsx")
                    break

        return achievements_msgs

    def get_profile(self, player, server):
        # Returns the rating and game stats of a given player in an embed
        wb = load_workbook("WarriorsvsSoldiers/database.xlsx")
        player_data = wb['Player records']
        ban_wb = load_workbook("WarriorsvsSoldiers/blacklist.xlsx")
        blacklist_players = ban_wb['Blacklist']
        
        for row in player_data:
            present = False
            if row[0].value == str(player.id):
                present = True
                player_sr = row[1].value
                stats = {'wins': row[2].value, 'total': row[3].value, 'soldier wins': row[4].value, 'soldier games': row[5].value, 'warrior wins': row[6].value, 'warrior games': row[7].value, 
                'win as soldier': row[8].value, 'played as soldier': row[9].value, 'win as warrior': row[10].value, 'played as warrior': row[11].value, 'win as coord': row[12].value, 'played as coord': row[13].value, 
                'win as queen': row[14].value, 'played as queen': row[15].value, 'win as warchief': row[16].value, 'played as warchief': row[17].value, 'win as ymir': row[20].value, 'played as ymir': row[21].value, 
                'win as false king': row[22].value, 'played as false king': row[23].value, 'win as ackerman': row[24].value, 'played as ackerman': row[25].value, 
                'win as mike': row[26].value, 'played as mike': row[27].value, 'win as scout': row[28].value, 'played as scout': row[29].value, 'win as spy': row[30].value, 'played as spy': row[31].value}
                break
        if present == False:
            # Add player data into player records
            player_sr = 1500
            player_data.append([str(player.id), 1500] + 30 * [0])
            stats = {'wins': 0, 'total': 0, 'soldier wins': 0, 'soldier games': 0, 'warrior wins': 0, 'warrior games': 0, 'win as soldier': 0, 'played as soldier': 0, 'win as warrior': 0, 'played as warrior': 0, 
            'win as coord': 0, 'played as coord': 0, 'win as queen': 0, 'played as queen': 0, 'win as warchief': 0, 'played as warchief': 0, 'win as ymir': 0, 'played as ymir':0, 'win as false king':0, 
            'played as false king': 0, 'win as ackerman': 0, 'played as ackerman': 0, 'win as mike': 0, 'played as mike': 0, 'win as scout': 0, 'played as scout': 0, 'win as spy': 0, 'played as spy': 0}
            wb.save("WarriorsvsSoldiers/database.xlsx")

        # Get total players in server
        player_rankings = [] # [[player id, sr], [...], ...]
        # Put all players into a list
        for row in player_data:
            if row[0].value == 'ID':
                continue
            else:
                player_rankings.append([row[0].value, row[1].value])
        
        # Sort by Exp from biggest to smallest
        player_rankings.sort(key = lambda x: x[1], reverse = True)

        # Get players in server
        server_users = list(map(lambda y: str(y.id), server.members))
        server_players = list(filter(lambda x:x[0] in server_users, player_rankings))
        total_players = len(server_players)

        # Get ranking
        i = 1
        for person in server_players:
            if person[0] == str(player.id):
                rank = i
                break
            else:
                i += 1

        # if '🎙 WvS Hosts' in [role.name for role in player.roles]:
        #     profile_title = '🎙 ' + player.name + '\'s Profile' + ' 🎙'
        if str(player.id) in [row[0].value for row in blacklist_players]:
            profile_title = '❌(Blacklisted) ' + player.name + '\'s Profile ❌'
        else:
            profile_title = player.name + '\'s Profile'

        profile = discord.Embed(title = profile_title, description = 'Skill Rating (SR): **' + str(int(round(player_sr, 0))) + '**', colour=0x0013B4)
        profile.set_thumbnail(url = player.avatar_url)

        profile.add_field(name = 'Rank', value = str(rank) + '/' + str(total_players), inline = False)
        profile.add_field(name = 'Games played', value = str(stats['total']))
        won_games = str(stats['wins'])
        if stats['total'] > 0:
            won_games += ' (' + str(round(100 * stats['wins'] / stats['total'], 1)) + '%)'
        profile.add_field(name = 'Games won', value = won_games)

        won_soldiers = '🛡 Soldiers: ' + str(stats['soldier wins']) + ' wins' + ' / ' + str(stats['soldier games']) + ' games'
        if stats['soldier games'] > 0:
            won_soldiers += ' **(' + str(round(100 * stats['soldier wins'] / stats['soldier games'], 1)) + '%)**'
        won_soldiers += ' 🛡'

        soldier_wins = 'Soldier: ' + str(stats['win as soldier']) + '/' + str(stats['played as soldier'])
        if stats['played as soldier'] > 0:
            soldier_wins += ' (' + str(round(100 * stats['win as soldier'] / stats['played as soldier'], 1)) + '%)'

        soldier_wins += '\nQueen: ' + str(stats['win as queen']) + '/' + str(stats['played as queen'])
        if stats['played as queen'] > 0:
            soldier_wins += ' (' + str(round(100 * stats['win as queen'] / stats['played as queen'], 1)) + '%)'

        soldier_wins += '\nAckerman: ' + str(stats['win as ackerman']) + '/' + str(stats['played as ackerman'])
        if stats['played as ackerman'] > 0:
            soldier_wins += ' (' + str(round(100 * stats['win as ackerman'] / stats['played as ackerman'], 1)) + '%)'

        soldier_wins += '\nMike Zacharias: ' + str(stats['win as mike']) + '/' + str(stats['played as mike'])
        if stats['played as mike'] > 0:
            soldier_wins += ' (' + str(round(100 * stats['win as mike'] / stats['played as mike'], 1)) + '%)'

        soldier_wins += '\nScout: ' + str(stats['win as scout']) + '/' + str(stats['played as scout'])
        if stats['played as scout'] > 0:
            soldier_wins += ' (' + str(round(100 * stats['win as scout'] / stats['played as scout'], 1)) + '%)'

        soldier_wins += '\nCoordinate: ' + str(stats['win as coord']) + '/' + str(stats['played as coord'])
        if stats['played as coord'] > 0:
            soldier_wins += ' (' + str(round(100 * stats['win as coord'] / stats['played as coord'], 1)) + '%)'

        won_warriors = '⚔ Warriors: ' + str(stats['warrior wins']) + ' wins' + ' / ' + str(stats['warrior games']) + ' games'
        if stats['warrior games'] > 0:
            won_warriors += ' **(' + str(round(100 * stats['warrior wins'] / stats['warrior games'], 1)) + '%)**'
        won_warriors += ' ⚔'

        warrior_wins = '\nWarrior: ' + str(stats['win as warrior']) + '/' + str(stats['played as warrior'])
        if stats['played as warrior'] > 0:
            warrior_wins += ' (' + str(round(100 * stats['win as warrior'] / stats['played as warrior'], 1)) + '%)'

        warrior_wins += '\nWarchief: ' + str(stats['win as warchief']) + '/' + str(stats['played as warchief'])
        if stats['played as warchief'] > 0:
            warrior_wins += ' (' + str(round(100 * stats['win as warchief'] / stats['played as warchief'], 1)) + '%)'

        warrior_wins += '\nFalse King: ' + str(stats['win as false king']) + '/' + str(stats['played as false king'])
        if stats['played as false king'] > 0:
            warrior_wins += ' (' + str(round(100 * stats['win as false king'] / stats['played as false king'], 1)) + '%)'
        
        warrior_wins += '\nYmir: ' + str(stats['win as ymir']) + '/' + str(stats['played as ymir'])
        if stats['played as ymir'] > 0:
            warrior_wins += ' (' + str(round(100 * stats['win as ymir'] / stats['played as ymir'], 1)) + '%)'
        
        warrior_wins += '\nSpy: ' + str(stats['win as spy']) + '/' + str(stats['played as spy'])
        if stats['played as spy'] > 0:
            warrior_wins += ' (' + str(round(100 * stats['win as spy'] / stats['played as spy'], 1)) + '%)'

        profile.add_field(name = won_soldiers, value = soldier_wins, inline = False)
        profile.add_field(name = won_warriors, value = warrior_wins, inline = False)
        
        return profile

    def get_achievements(self, player):
        # Returns profile of player in an embed
        wb = load_workbook("WarriorsvsSoldiers/database.xlsx")
        player_data = wb['Player records']

        # Extract data
        inside = False
        for row in player_data:
            if row[0].value == str(player.id):
                # [Name, No. games played, Current win streak, Max win streak, Soldier wins, Warrior wins, Coordinate wins, Queen/Ackerman/Mike/Scout wins, Warchief/False King/Ymir/Spy wins]
                badges = [player.name, row[3].value, row[18].value, row[19].value, row[8].value, row[10].value, row[12].value, 
                (row[14].value + row[24].value + row[26].value + row[28].value), (row[16].value + row[22].value + row[20].value + row[30].value)]
                inside = True
                break
        if inside == False:
            badges = [player.name, 0, 0, 0, 0, 0, 0, 0, 0]

        # Put info into embed
        num_badges = 0

        number_games_played = badges[1]
        number_games_badges = ''
        number_games_next = str(number_games_played)
        for step in sorted(list(self.number_games_achievements)):
            if number_games_played < step:
                number_games_next = str(number_games_played) + '/' + str(step)
                break
            number_games_badges += self.badge_emojis[self.number_games_achievements[step]]
            num_badges += 1

        cur_streak = badges[2]
        max_streak = badges[3]
        consecutive_wins_badges = ''
        consecutive_next = str(max_streak)
        for step in sorted(list(self.consecutive_wins_achievements)):
            if max_streak < step:
                consecutive_next = str(cur_streak) + '/' + str(step)
                break
            consecutive_wins_badges += self.badge_emojis[self.consecutive_wins_achievements[step]]
            num_badges += 1

        soldier_wins = badges[4]
        soldier_wins_badges = ''
        soldier_wins_next = str(soldier_wins)
        for step in sorted(list(self.soldier_wins_achievements)):
            if soldier_wins < step:
                soldier_wins_next = str(soldier_wins) + '/' + str(step)
                break
            soldier_wins_badges += self.badge_emojis[self.soldier_wins_achievements[step]]
            num_badges += 1

        warrior_wins = badges[5]
        warrior_wins_badges = ''
        warrior_wins_next = str(warrior_wins)
        for step in sorted(list(self.warrior_wins_achievements)):
            if warrior_wins < step:
                warrior_wins_next = str(warrior_wins) + '/' + str(step)
                break
            warrior_wins_badges += self.badge_emojis[self.warrior_wins_achievements[step]]
            num_badges += 1

        coordinate_wins = badges[6]
        coordinate_wins_badges = ''
        coordinate_wins_next = str(coordinate_wins)
        for step in sorted(list(self.coordinate_wins_achievements)):
            if coordinate_wins < step:
                coordinate_wins_next = str(coordinate_wins) + '/' + str(step)
                break
            coordinate_wins_badges += self.badge_emojis[self.coordinate_wins_achievements[step]]
            num_badges += 1

        queen_wins = badges[7]
        queen_wins_badges = ''
        queen_wins_next = str(queen_wins)
        for step in sorted(list(self.queen_wins_achievements)):
            if queen_wins < step:
                queen_wins_next = str(queen_wins) + '/' + str(step)
                break
            queen_wins_badges += self.badge_emojis[self.queen_wins_achievements[step]]
            num_badges += 1

        warchief_wins = badges[8]
        warchief_wins_badges = ''
        warchief_wins_next = str(warchief_wins)
        for step in sorted(list(self.warchief_wins_achievements)):
            if warchief_wins < step:
                warchief_wins_next = str(warchief_wins) + '/' + str(step)
                break
            warchief_wins_badges += self.badge_emojis[self.warchief_wins_achievements[step]]
            num_badges += 1

        player_achievements = discord.Embed(title = 'Badges for Warriors vs Soldiers', description = 'Number of badges collected: **' + str(num_badges) + '**', colour = 0xC0C0C0)
        player_achievements.set_author(name = player.name, icon_url = player.avatar_url)
        player_achievements.set_thumbnail(url = player.avatar_url)

        player_achievements.add_field(name = 'Number of games played', value = number_games_next)
        player_achievements.add_field(name = 'Badges', value = number_games_badges if number_games_badges != '' else '-')
        player_achievements.add_field(name = '⠀', value='⠀')
        player_achievements.add_field(name = 'Number of wins in a row', value = consecutive_next)
        player_achievements.add_field(name = 'Badges', value = consecutive_wins_badges if consecutive_wins_badges != '' else '-')
        player_achievements.add_field(name = '⠀', value='⠀')
        player_achievements.add_field(name = 'Number of wins as Soldier', value = soldier_wins_next)
        player_achievements.add_field(name = 'Badges', value = soldier_wins_badges if soldier_wins_badges != '' else '-')
        player_achievements.add_field(name = '⠀', value='⠀')
        player_achievements.add_field(name = 'Number of wins as Warrior', value = warrior_wins_next)
        player_achievements.add_field(name = 'Badges', value = warrior_wins_badges if warrior_wins_badges != '' else '-')
        player_achievements.add_field(name = '⠀', value='⠀')
        player_achievements.add_field(name = 'Number of wins as Coordinate', value = coordinate_wins_next)
        player_achievements.add_field(name = 'Badges', value = coordinate_wins_badges if coordinate_wins_badges != '' else '-')
        player_achievements.add_field(name = '⠀', value='⠀')
        player_achievements.add_field(name = 'Number of wins as Optional Soldier role', value = queen_wins_next)
        player_achievements.add_field(name = 'Badges', value = queen_wins_badges if queen_wins_badges != '' else '-')
        player_achievements.add_field(name = '⠀', value='⠀')
        player_achievements.add_field(name = 'Number of wins as Optional Warrior role', value = warchief_wins_next)
        player_achievements.add_field(name = 'Badges', value = warchief_wins_badges if warchief_wins_badges != '' else '-')
        player_achievements.add_field(name = '⠀', value='⠀')

        return player_achievements

    def get_game_stats(self, page_no):
        # Returns past game statistics
        wb = load_workbook("WarriorsvsSoldiers/database.xlsx")
        game_data = wb['Game records']
        player_data = wb['Player records']

        page_descriptions = {1: 'Game Stats', 2: 'Role Stats'}

        game_stats = discord.Embed(title = '⚔ Warriors vs Soldiers Stats 🛡 ', description = 'Page ' + str(page_no) + ': ' + page_descriptions[page_no], colour=0x0013B4)

        if page_no == 1:
            # Game stats
            games_played = 'Soldiers: ' + str(game_data['H2'].value)
            if game_data['H10'].value > 0:
                games_played += ' (' + str(round(100 * game_data['H2'].value / game_data['H10'].value, 1)) + '%)'
            games_played += '\nWarriors: ' + str(game_data['H8'].value)
            if game_data['H10'].value > 0:
                games_played += ' (' + str(round(100 * game_data['H8'].value / game_data['H10'].value, 1)) + '%)'
            game_stats.add_field(name = 'Total games played: ' + str(game_data['H10'].value), value = games_played, inline = False)

            players5 = 'Soldiers: ' + str(game_data['B2'].value)
            if game_data['B10'].value > 0:
                players5 += ' (' + str(round(100 * game_data['B2'].value / game_data['B10'].value, 1)) + '%)'
            players5 += '\nWarriors: ' + str(game_data['B8'].value)
            if game_data['B10'].value > 0:
                players5 += ' (' + str(round(100 * game_data['B8'].value / game_data['B10'].value, 1)) + '%)'
            game_stats.add_field(name = '5 Players: ' + str(game_data['B10'].value), value = players5)

            players6 = 'Soldiers: ' + str(game_data['C2'].value)
            if game_data['C10'].value > 0:
                players6 += ' (' + str(round(100 * game_data['C2'].value / game_data['C10'].value, 1)) + '%)'
            players6 += '\nWarriors: ' + str(game_data['C8'].value)
            if game_data['C10'].value > 0:
                players6 += ' (' + str(round(100 * game_data['C8'].value / game_data['C10'].value, 1)) + '%)'
            game_stats.add_field(name = '6 Players: ' + str(game_data['C10'].value), value = players6)

            players7 = 'Soldiers: ' + str(game_data['D2'].value)
            if game_data['D10'].value > 0:
                players7 += ' (' + str(round(100 * game_data['D2'].value / game_data['D10'].value, 1)) + '%)'
            players7 += '\nWarriors: ' + str(game_data['D8'].value)
            if game_data['D10'].value > 0:
                players7 += ' (' + str(round(100 * game_data['D8'].value / game_data['D10'].value, 1)) + '%)'
            game_stats.add_field(name = '7 Players: ' + str(game_data['D10'].value), value = players7)

            players8 = 'Soldiers: ' + str(game_data['E2'].value)
            if game_data['E10'].value > 0:
                players8 += ' (' + str(round(100 * game_data['E2'].value / game_data['E10'].value, 1)) + '%)'
            players8 += '\nWarriors: ' + str(game_data['E8'].value)
            if game_data['E10'].value > 0:
                players8 += ' (' + str(round(100 * game_data['E8'].value / game_data['E10'].value, 1)) + '%)'
            game_stats.add_field(name = '8 Players: ' + str(game_data['E10'].value), value = players8)

            players9 = 'Soldiers: ' + str(game_data['F2'].value)
            if game_data['F10'].value > 0:
                players9 += ' (' + str(round(100 * game_data['F2'].value / game_data['F10'].value, 1)) + '%)'
            players9 += '\nWarriors: ' + str(game_data['F8'].value)
            if game_data['F10'].value > 0:
                players9 += ' (' + str(round(100 * game_data['F8'].value / game_data['F10'].value, 1)) + '%)'
            game_stats.add_field(name = '9 Players: ' + str(game_data['F10'].value), value = players9)

            players10 = 'Soldiers: ' + str(game_data['G2'].value)
            if game_data['G10'].value > 0:
                players10 += ' (' + str(round(100 * game_data['G2'].value / game_data['G10'].value, 1)) + '%)'
            players10 += '\nWarriors: ' + str(game_data['G8'].value)
            if game_data['G10'].value > 0:
                players10 += ' (' + str(round(100 * game_data['G8'].value / game_data['G10'].value, 1)) + '%)'
            game_stats.add_field(name = '10 Players: ' + str(game_data['G10'].value), value = players10)

        elif page_no == 2:
            # Role stats
            soldier_wins, soldier_games = 0, 0
            warrior_wins, warrior_games = 0, 0
            coordinate_wins, coordinate_games = 0, 0
            queen_wins, queen_games = 0, 0
            warchief_wins, warchief_games = 0, 0
            ymir_wins, ymir_games = 0, 0
            falseking_wins, falseking_games = 0, 0
            ackerman_wins, ackerman_games = 0, 0
            mike_wins, mike_games = 0, 0
            scout_wins, scout_games = 0, 0
            spy_wins, spy_games = 0, 0

            for i in range(2, player_data.max_row + 1):
                soldier_wins += player_data['I' + str(i)].value
                soldier_games += player_data['J' + str(i)].value
                warrior_wins += player_data['K' + str(i)].value
                warrior_games += player_data['L' + str(i)].value
                coordinate_wins += player_data['M' + str(i)].value
                coordinate_games += player_data['N' + str(i)].value
                queen_wins += player_data['O' + str(i)].value
                queen_games += player_data['P' + str(i)].value
                warchief_wins += player_data['Q' + str(i)].value
                warchief_games += player_data['R' + str(i)].value
                ymir_wins += player_data['U' + str(i)].value
                ymir_games += player_data['V' + str(i)].value
                falseking_wins += player_data['W' + str(i)].value
                falseking_games += player_data['X' + str(i)].value
                ackerman_wins += player_data['Y' + str(i)].value
                ackerman_games += player_data['Z' + str(i)].value
                mike_wins += player_data['AA' + str(i)].value
                mike_games += player_data['AB' + str(i)].value
                scout_wins += player_data['AC' + str(i)].value
                scout_games += player_data['AD' + str(i)].value
                spy_wins += player_data['AE' + str(i)].value
                spy_games += player_data['AF' + str(i)].value

            soldier_stats = 'Games: ' + str(soldier_games)
            soldier_stats += '\nWins: ' + str(soldier_wins)
            if soldier_wins > 0:
                soldier_stats += ' (' + str(round(100 * soldier_wins / soldier_games, 1)) + '%)'
            game_stats.add_field(name = '🛡Soldier🛡', value = soldier_stats)

            warrior_stats = 'Games: ' + str(warrior_games)
            warrior_stats += '\nWins: ' + str(warrior_wins)
            if warrior_wins > 0:
                warrior_stats += ' (' + str(round(100 * warrior_wins / warrior_games, 1)) + '%)'
            game_stats.add_field(name = '⚔Warrior⚔', value = warrior_stats)

            coordinate_stats = 'Games: ' + str(coordinate_games)
            coordinate_stats += '\nWins: ' + str(coordinate_wins)
            if coordinate_wins > 0:
                coordinate_stats += ' (' + str(round(100 * coordinate_wins / coordinate_games, 1)) + '%)'
            game_stats.add_field(name = '🗺Coordinate🗺', value = coordinate_stats)

            queen_stats = 'Games: ' + str(queen_games)
            queen_stats += '\nWins: ' + str(queen_wins)
            if queen_wins > 0:
                queen_stats += ' (' + str(round(100 * queen_wins / queen_games, 1)) + '%)'
            game_stats.add_field(name = '👼Queen👼', value = queen_stats)

            warchief_stats = 'Games: ' + str(warchief_games)
            warchief_stats += '\nWins: ' + str(warchief_wins)
            if warchief_wins > 0:
                warchief_stats += ' (' + str(round(100 * warchief_wins / warchief_games, 1)) + '%)'
            game_stats.add_field(name = '🦹‍♂️Warchief🦹‍♂️', value = warchief_stats)

            ymir_stats = 'Games: ' + str(ymir_games)
            ymir_stats += '\nWins: ' + str(ymir_wins)
            if ymir_wins > 0:
                ymir_stats += ' (' + str(round(100 * ymir_wins / ymir_games, 1)) + '%)'
            game_stats.add_field(name = '🤷‍♀️Ymir🤷‍♀️', value = ymir_stats)

            falseking_stats = 'Games: ' + str(falseking_games)
            falseking_stats += '\nWins: ' + str(falseking_wins)
            if falseking_wins > 0:
                falseking_stats += ' (' + str(round(100 * falseking_wins / falseking_games, 1)) + '%)'
            game_stats.add_field(name = '🕴False King🕴', value = falseking_stats)

            ackerman_stats = 'Games: ' + str(ackerman_games)
            ackerman_stats += '\nWins: ' + str(ackerman_wins)
            if ackerman_wins > 0:
                ackerman_stats += ' (' + str(round(100 * ackerman_wins / ackerman_games, 1)) + '%)'
            game_stats.add_field(name = '💂Ackerman💂', value = ackerman_stats)

            mike_stats = 'Games: ' + str(mike_games)
            mike_stats += '\nWins: ' + str(mike_wins)
            if mike_wins > 0:
                mike_stats += ' (' + str(round(100 * mike_wins / mike_games, 1)) + '%)'
            game_stats.add_field(name = '<:aotSmirk2:455223780957224961>Mike Zacharias <:aotSmirk2:455223780957224961>', value = mike_stats)

            scout_stats = 'Games: ' + str(scout_games)
            scout_stats += '\nWins: ' + str(scout_wins)
            if scout_wins > 0:
                scout_stats += ' (' + str(round(100 * scout_wins / scout_games, 1)) + '%)'
            game_stats.add_field(name = '🏇Scout🏇', value = scout_stats)

            spy_stats = 'Games: ' + str(spy_games)
            spy_stats += '\nWins: ' + str(spy_wins)
            if spy_wins > 0:
                spy_stats += ' (' + str(round(100 * spy_wins / spy_games, 1)) + '%)'
            game_stats.add_field(name = '🕵️‍♀️Spy🕵️‍♀️', value = spy_stats)

        return game_stats

    def get_leaderboard(self, server, page=1):
        # Returns the names of the top players, with 10 per page, in an embed
        wb = load_workbook("WarriorsvsSoldiers/database.xlsx")
        player_data = wb['Player records']
        
        player_rankings = {} # {player id: sr, ...}
        # Put all players into a dictionary
        for line in player_data:
            player_rankings[line[0].value] = line[1].value

        server_users = server.members
        server_players = []
        for user in server_users:
            if str(user.id) in player_rankings:
                server_players.append([user.mention, player_rankings[str(user.id)]])
        
        # Sort by SR from biggest to smallest
        server_players.sort(key = lambda x: x[1], reverse = True)

        # Top 10x players
        try: 
            page_no = int(page)
        except:
            page_no = 1

        if page_no < 1:
            page_no = 1

        # Normalize to number of server users
        page_no = min(math.ceil(len(server_players) / 10), page_no)

        # Put names into embed
        all_names = ''
        all_sr = ''
        nums = {1:'1⃣', 2:'2⃣', 3:'3⃣', 4:'4⃣', 5:'5⃣', 6:'6⃣', 7:'7⃣', 8:'8⃣', 9:'9⃣', 10:'🔟'}
        for rank in range((page_no - 1) * 10 + 1, min(len(server_players) + 1, page_no * 10 + 1)):
            if rank == 1:
                all_names += '🥇 '
            elif rank == 2:
                all_names += '🥈 '
            elif rank == 3:
                all_names += '🥉 '
            elif rank <= 10:
                all_names += str(nums[rank]) + ' '
            else:
                all_names += '#' + str(rank) + ' '
            all_names += server_players[rank-1][0] + '\n'
            if rank <= 10:
                all_sr += '🔹 ' + str(int(round(server_players[rank-1][1], 0))) + ' 🔹' + '\n'
            else:
                all_sr += ' ￶￵ ￶￵ ￶￵  ￶￵ ￶￵ ￶￵  ￶￵ ￶￵ ￶￵ ￶￵ '+ str(int(round(server_players[rank-1][1], 0))) + '\n'

        leaderboard = discord.Embed(title = 'Leaderboard for Warriors vs Soldiers', description = 'Page ' + str(page_no), colour=0x0013B4)
        # leaderboard.set_thumbnail(url = 'https://i.imgur.com/RZBR3Hb.png')

        leaderboard.add_field(name = 'Player', value = all_names)
        leaderboard.add_field(name = 'Skill Rating (SR)', value = all_sr)
        return leaderboard

    def get_commands(self, *command_query):
        soldier_info = '🛡**Soldier**🛡\n\n\
The most common role in the game, the Soldier represents the uninformed majority. Their role is find out who the traitors within the military are, reach the basement before they destroy all 3 Walls, \
and help protect the identity of the Coordinate.\n\n\
**Tip**: The Soldier is simultaneously the weakest and strongest role in the game. Even though the Soldier has no knowledge of any one else\'s identity, a Soldier has \
no need to conceal their identity and are free to act however they want.'
        warrior_info = '⚔**Warrior**⚔\n\n\
The Warrior is the spy within the ranks of the military. Their role is to avoid being discovered by the Soldiers, persuade them to allow you into their expeditions, and sabotage them. If that fails, \
they can still win by identifying the Coordinate within the ranks of the soldiers.\n\n\
**Tip**: The Warrior relies on deception and cunning to succeed. Teamwork and coordination with other Warriors is essential, to prevent your identities \
from being exposed. This must be done discreetly, however, as using DMs to communicate with other Warriors is against the rules and spirit of the game.'
        coordinate_info = '🗺**Coordinate**🗺\n\n\
The Coordinate holds the greatest power in the game, but with it comes a heavy responsibility. The Coordinate is aware of the Warriors\' identities from the start, but has to find \
a way to communicate this information to the Soldiers without being noticed by the Warriors, who can kidnap the Coordinate if the Soldiers reach the Basement.\n\n\
**Tip**: The Coordinate\'s playstyle can change depending on the situation. For example, if the other Soldiers are doing a good job identifying the Warriors, it would be advisable to focus on \
deflecting suspicion away from themselves instead.'
        queen_info = '👼**Queen**👼\n\n\
The Queen has the unique position of knowing the Coordinate\'s identity, which allows her to make better decisions during expeditions at no added risk. Her role is to protect \
the Coordinate using this information by acting like they would.\n\n\
**Tip**: While the Queen has the advantage of knowing the Coordinate\'s identity, she also has to be careful not to accidentally reveal the Coordinate\'s identity by favoring them on expeditions too much.'
        ackerman_info = '💂**Ackerman**💂\n\n\
The Ackerman has the ability to secure the Walls in **one** expedition that they\'re in. This prevents any Warriors on the expedition from destroying it. However, doing so also \
alerts any Warriors in the expedition of the Ackerman\'s identity, making it easier for them to identity the Coordinate.\n\n\
**Tip**: Knowing if and when to secure the Walls is key to playing an Ackerman well. For example, if the Soldiers are winning, it might be best to not secure the Walls at all, reducing the chances \
of the Warriors kidnapping the Coordinate at the end.'
        mike_info = '<:aotSmirk2:455223780957224961>**Mike Zacharias** <:aotSmirk2:455223780957224961>\n\n\
Humanity\'s 2nd strongest soldier **Mike Zacharias** has the incredible ability to sniff out Titans in expeditions <:aotSmirk2:455223780957224961>. \
During the approval phase of all the expeditions that he is in, Mike will be told how many Titans the expedition contains, if any. Both Warriors and the Coordinate are considered Titans.\n\n\
**Tip**: The knowledge of how many Titans are in an expedition can be very valuable to both Soldiers and Warriors. Be careful what you do with that information. \
Prove yourself worthy of the title of humanity\'s 2nd strongest soldier!'
        scout_info = '🏇**Scout**🏇\n\n\
The Scout is the guiding light of an expedition. If the Scout is in an expedition, they will automatically fire a signal flare, alerting everyone of their presence in the expedition.\n\n\
**Tip**: The knowledge of the Scout\'s identity can benefit both sides in the game. The Scout should try to help the Soldiers identify them correctly, while working with their fellow Soldiers \
to confuse the Warriors as to who they are.'
        falseking_info = '🕴**False King**🕴\n\n\
The False King appears as the Coordinate to the Queen, in addition to the real Coordinate. This gives him the ability to confuse her by acting as the Coordinate, which can \
help turn the tide in the Warriors\' favor.\n\n\
**Tip**: As the False King, it can be especially advantageous to succeed expeditions which you are in, solely to gain the Queen\'s trust.'
        warchief_info = '🦹‍♂️**Warchief**🦹‍♂️\n\n\
The Warchief has the ability to conceal his identity from the Coordinate. This allows more possibilities to sow confusion within the Soldiers\' ranks without worrying about the \
Coordinate stepping in.\n\n\
**Tip**: The Warchief\'s role acts like a double-edged blade. Without the Coordinate\'s knowing the Warchief\'s identity, he can more easily gain the Soldiers\' trust. \
However, the very fact that the Coordinate is unaware of the Warchief\'s identity can make it harder for the Warriors to identify the Coordinate.'
        ymir_info = '🤷‍♀️**Ymir**🤷‍♀️\n\n\
Ymir is also on the Warriors\' side, but does not know of their identities, nor do they know hers. To achieve success, she must keep a close watch \
for potential comrades during expeditions.\n\n\
**Tip**: In smaller expeditions, it can be risky to choose to destroy the Walls, as there might be another Warrior in your group which would give both of you away. \
It is usually a good idea to wait until the later (and larger) expeditions to make your move.'
        spy_info = '🕵️‍♀️**Spy**🕵️‍♀️\n\n\
A highly skilled infiltrator, the Spy has the ability to flip the votes during the approval phase of an expedition. \
She may only do this once, however, as everyone will be alerted that the votes have been flipped.\n\n\
**Tip**: The Spy needs to be strategic about when to flip the votes to favor the Warriors, as it can easily backfire if people do not vote the way you expect.'
        blessing_info = '🔮**Ymir\'s Blessing**🔮\n\n\
Ymir\'s Blessing allows a player who has it to check another player\'s true allegiance (Warriors or Soldiers). At the start of the 3rd expedition, a random player will be given Ymir\'s blessing. \
They may use it on another player to find out their allegiance (will be DMed to them in private).\n\n\
Ymir\'s blessing will then be passed onto the player who was investigated, who can use it at the start of the next expedition, and so on. \
However, anyone who has had Ymir\'s blessing previously is granted immunity to being investigated by future holders of it.'
        paths_info = '📢**Paths**📢\n\n\
The Paths ability allows players to make an announcement anonymously to everyone in the game. When activated, a random person will be designated as the Paths holder at the start of every expedition. \
They will then be able to send a single message to the game channel without revealing their role or identity.'

        if command_query:
            commands_dict = {'host':'Creates a new lobby with you as the host. Add \'casual\' to the command to host an unranked game.',
                            'join':'Joins an existing lobby.',
                            'leave':'Leave your current lobby',
                            'kick':'Removes a player from the lobby (Only the host can kick)',
                            'start':'Starts the game (requires 5-10 players)',
                            'reset':'Stops the existing game or clears the current lobby (only the host can reset once the game starts)',
                            'add':'Adds the specified optional role to the game. (E.g. ~add queen)',
                            'remove':'Removes the specified optional role from the game. (E.g. ~add queen)',
                            'randomroles': 'Toggles randomization of optional roles when starting a game.',
                            'players':'If game hasn\'t started yet: Brings up the current list of players in the lobby.\n\
If game has started: Brings up the current list of players, arranged in order of their position in the queue to be Commander.',
                            'next':'Starts the next expedition. Used by the host after the previous expedition has ended.',
                            'pick':'The command used by the current Commander to select the expedition team (E.g. ~pick @armin)',
                            'kidnap': 'Usable if the Soldiers reach the basement, any Warrior can use this command to pick who they think is the Coordinate (E.g. ~kidnap @eren)',
                            'status':'Brings up the current game status. This includes the number of players, the number of Warriors and Soldiers, the current progress towards the Basement \
and status of the Walls, the results of previous expeditions and information on the current expedition.',
                            'tutorial':'Provides a rundown of how to play the game, with detailed information about each role.',
                            'profile':'Calls up the profile of a given person (E.g. ~profile @levi).',
                            'badges':'Checks the badges a given user has. Use just ~badges to check your own badges.',
                            'gamestats':'Brings up the past records of all games played.',
                            'leaderboard/lb': 'Brings up the leaderboard, listing the top 10 players on the server. Add a number behind to see subsequent pages (e.g. ~lb 2).',
                            'role': 'Checks your role while in a game.',
                            'roles': 'Shows the list of roles currently in the game.'}

            roles_dict = {'soldier': soldier_info,
                        'warrior': warrior_info,
                        'coordinate': coordinate_info,
                        'queen': queen_info,
                        'ackerman': ackerman_info,
                        'mike': mike_info,
                        'scout': scout_info,
                        'falseking': falseking_info,
                        'warchief': warchief_info,
                        'ymir': ymir_info,
                        'spy': spy_info,
                        'blessing': blessing_info,
                        'paths': paths_info}

            if command_query[0] in commands_dict:
                return discord.Embed(title = '~' + command_query[0], description = commands_dict[command_query[0]], colour = 0x0013B4)
            elif command_query[0] in roles_dict:
                return discord.Embed(description = roles_dict[command_query[0]], colour = 0x0013B4)

        else:
            lobby_commands = '~host\n\
~join\n\
~leave\n\
~kick <@person>\n\
~start\n\
~reset\n\
~add <role>\n\
~remove <role>\n\
~randomroles\n\
~players\n\
~tutorial\n\
~profile <@person>\n\
~badges <@person>\n\
~gamestats\n\
~leaderboard'

            game_commands = '~role\n\
~roles\n\
~next\n\
~pick <@person>\n\
~kidnap <@person>\n\
~status\n\
~players'

            commands = discord.Embed(title = 'List of commands for Warriors vs Soldiers', description = 'For more information about a specific command, type ~help <command>. (E.g. ~help kidnap)', colour=0x0013B4)
            commands.add_field(name = 'Lobby commands', value = lobby_commands)
            commands.add_field(name = 'In-game commands', value = game_commands)
            return commands

    def tutorial(self, num):
        if num == 1:
            tutorial1 = 'Warriors vs Soldiers is an Attack on Titan themed party game modelling a conflict between two groups: an informed minority (the Warriors), \
and an uninformed majority (the Soldiers).\n\n\
**📌Rules of the game📌**\n\n\
:one: The objective of the Soldiers is to figure out who the Warriors are, while the Warriors aim to conceal their identity from the Soldiers.\n\n\
:two: The team that completes their objectives first win (succeed 3 expeditions for Soldiers, destroy all 3 Walls for Warriors).\n\n\
:three: Messaging another player in private (DM) is not allowed while in a game.\n\n\
:four: Revealing your role via screenshot is also not allowed.\n\n\
:five: To invite people to join, you can ping `@Warriors vs. Soldiers` (do not abuse pings, and do not ping other roles or random people).\n\n\
:six: Be civil and respectful to others when playing, keeping in mind these rules and the server <#458311147234263051>. Failure to do so may result in you getting blacklisted.\n\n\
Click on ▶ below for more information about the game.'
            return tutorial1

        if num == 2:
            tutorial2 = '**📖Backstory📖**\n\n\
The Year 850 was a pivotal year in the world of Attack on Titan. Around this time, it was discovered that Titan shifters who call themselves "Warriors" have infiltrated the Walls, \
and are attempting to wipe out the remainder of humanity by either destroying the Walls with their Titan powers or capturing the Coordinate, an unimaginable source of power.\n\n\
At approximately the same time, the Survey Corps, soldiers who have dedicated their lives to the survival of humanity, are on the verge of a major breakthrough. They have discovered \
that a certain basement in the border town of Shiganshina contains valuable secrets about the Titans, secrets that could allow them to overcome the Titan menace once and for all. \
Acting on that information, the Survey Corps prepare to embark on a series of expeditions to reach the Basement and secure this secret.\n\n\
This game puts you and your friends at the centre of this conflict. In a race against time, whichever side manages to reach their objectives first will decide the fate of humanity.\n\n\
Click on ▶ below to continue.'
            return tutorial2

        elif num == 3:
            tutorial3 = '**🎲Gameplay🎲**\n\n\
At the start of the game, each player is secretly assigned a role (through DMs) affliated with one of these teams. The Warriors are made aware of each other without the Soldiers knowing - \
the only thing the Soldiers know is how many Warriors exist, not who they are. Additionally, the player with the Coordinate role is also made aware of the Warriors\' identities.\n\n\
The game consists of up to 5 expeditions, each expedition consisting of 3 phases to decide its success or failure.\n\n\
**🔹Planning Phase🔹**\n\n\
At the start of this phase, one of the players (either a Soldier or Warrior) is selected, based on a randomized queue implemented at the start of the game, to be the Commander. \
The Commander selects a certain number of players to form the expedition team (the Commander may choose to join the expedition team themselves).\n\n\
**🔹Approval Phase🔹**\n\n\
After the Commander has picked the expedition team, all of the players discuss the Commander\'s choice, and vote in private (via DMs) on whether to accept the team make-up or not. \
If a majority of players votes no to the proposal or if it\'s a tie, the Commander position passes on to the next player in line, who proposes their own expedition team. This continues \
until a majority of players agree with the current Commander\'s expedition team, at which point the round moves to the action phase.'
            return tutorial3

        elif num == 4:
            tutorial4 = '**🔹Action Phase🔹**\n\n\
Once an expedition team is agreed on, the players in the expedition team then "embark" on the expedition. A Soldier must choose to help the expedition succeed, while a Warrior may either \
help the expedition succeed as well, or secretly sabotage the expedition (via DM), destroying one of the Walls in the process.\n\n\
Once everyone has made their decision, the individual choices are \
shuffled then revealed to everyone, minus the names of the players who made the choice. If all the players choose to succeed the expedition, the Soldiers advance one step closer to the Basement. \
But if a Warrior sabotages the expedition, the outermost Wall is destroyed and the Warriors move one step closer to victory.\n\n\
This process then repeats until one side has accumulated 3 victories (Soldiers reach the Basement after 3 expeditions, Warriors have 3 walls to destroy).\n\n\
**🏆Win conditions and the Coordinate🏆**\n\n\
If the Warriors manage to destroy all 3 Walls before the Soldiers reach the basement, then Warriors win.\n\n\
If the Soldiers manage to reach the Basement first, the Warriors have one last chance to win by correctly identifying and kidnapping the Coordinate, who knows their identities. \
If they fail to do so, then Soldiers win.\n\n\
Click on ▶ below to see the detailed descriptions of each role.'
            return tutorial4

        elif num == 5:
            tutorial5 = '**Basic Roles**\n\n\
🛡**Soldier**🛡\n\n\
The most common role in the game, the Soldier represents the uninformed majority. Their role is find out who the traitors within the military are, reach the basement before they destroy all 3 Walls, \
and help protect the identity of the Coordinate.\n\n\
Tip: The Soldier is simultaneously the weakest and strongest role in the game. Even though the Soldier has no knowledge of any one else\'s identity, a Soldier has \
no need to conceal their identity and are free to act however they want.\n\n\
⚔**Warrior**⚔\n\n\
The Warrior is the spy within the ranks of the military. Their role is to avoid being discovered by the Soldiers, persuade them to allow you into their expeditions, and sabotage them. If that fails, \
they can still win by identifying the Coordinate within the ranks of the soldiers.\n\n\
Tip: The Warrior relies on deception and cunning to succeed. Teamwork and coordination with other Warriors is essential, to prevent your identities \
from being exposed. This must be done discreetly, however, as using DMs to communicate with other Warriors is against the rules and spirit of the game.\n\n\
🗺**Coordinate**🗺\n\n\
The Coordinate holds the greatest power in the game, but with it comes a heavy responsibility. The Coordinate is aware of the Warriors\' identities from the start, but has to find \
a way to communicate this information to the Soldiers without being noticed by the Warriors, who can kidnap the Coordinate if the Soldiers reach the Basement.\n\n\
Tip: The Coordinate\'s playstyle can change depending on the situation. For example, if the other Soldiers are doing a good job identifying the Warriors, it would be advisable to focus on \
deflecting suspicion away from themselves instead.'
            return tutorial5

        elif num == 6:
            tutorial6 = '**Optional Soldier Roles (1)**\n\n\
👼**Queen**👼\n\n\
The Queen has the unique position of knowing the Coordinate\'s identity, which allows her to make better decisions during expeditions at no added risk. Her role is to protect \
the Coordinate using this information by acting like they would.\n\n\
Tip: While the Queen has the advantage of knowing the Coordinate\'s identity, she also has to be careful not to accidentally reveal the Coordinate\'s identity by favoring them on expeditions too much.\n\n\
💂**Ackerman**💂\n\n\
The Ackerman has the ability to secure the Walls in **one** expedition that they\'re in. This prevents any Warriors on the expedition from destroying it. However, doing so also \
alerts any Warriors in the expedition of the Ackerman\'s identity, making it easier for them to identity the Coordinate.\n\n\
Tip: Knowing if and when to secure the Walls is key to playing an Ackerman well. For example, if the Soldiers are winning, it might be best to not secure the Walls at all, reducing the chances \
of the Warriors kidnapping the Coordinate at the end.'
            return tutorial6

        elif num == 7:
            tutorial7 = '**Optional Soldier Roles (2)**\n\n\
<:aotSmirk2:455223780957224961>**Mike Zacharias** <:aotSmirk2:455223780957224961>\n\n\
Humanity\'s 2nd strongest soldier **Mike Zacharias** has the incredible ability to sniff out Titans in expeditions <:aotSmirk2:455223780957224961>. \
During the approval phase of all the expeditions that he is in, Mike will be told how many Titans the expedition contains, if any. Both Warriors and the Coordinate are considered Titans.\n\n\
Tip: The knowledge of how many Titans are in an expedition can be very valuable to both Soldiers and Warriors. Be careful what you do with that information. \
Prove yourself worthy of the title of humanity\'s 2nd strongest soldier!\n\n\
🏇**Scout**🏇\n\n\
The Scout is the guiding light of an expedition. If the Scout is in an expedition, they will automatically fire a signal flare, alerting everyone of their presence in the expedition.\n\n\
Tip: The knowledge of the Scout\'s identity can benefit both sides in the game. The Scout should try to help the Soldiers identify them correctly, while working with their fellow Soldiers \
to confuse the Warriors as to who they are.'
            return tutorial7

        elif num == 8:
            tutorial8 = '**Optional Warrior Roles (1)**\n\n\
🕴**False King**🕴\n\n\
The False King appears as the Coordinate to the Queen, in addition to the real Coordinate. This gives him the ability to confuse her by acting as the Coordinate, which can \
help turn the tide in the Warriors\' favor.\n\n\
Tip: As the False King, it can be especially advantageous to succeed expeditions which you are in, solely to gain the Queen\'s trust.\n\n\
🦹‍♂️**Warchief**🦹‍♂️\n\n\
The Warchief has the ability to conceal his identity from the Coordinate. This allows more possibilities to sow confusion within the Soldiers\' ranks without worrying about the \
Coordinate stepping in.\n\n\
Tip: The Warchief\'s role acts like a double-edged blade. Without the Coordinate\'s knowing the Warchief\'s identity, he can more easily gain the Soldiers\' trust. \
However, the very fact that the Coordinate is unaware of the Warchief\'s identity can make it harder for the Warriors to identify the Coordinate.'
            return tutorial8

        elif num == 9:
            tutorial9 = '**Optional Warrior Roles (2)**\n\n\
🤷‍♀️**Ymir**🤷‍♀️\n\n\
Ymir is also on the Warriors\' side, but does not know of their identities, nor do they know hers. To achieve success, she must keep a close watch \
for potential comrades during expeditions.\n\n\
Tip: In smaller expeditions, it can be risky to choose to destroy the Walls, as there might be another Warrior in your group which would give both of you away. \
It is usually a good idea to wait until the later (and larger) expeditions to make your move.\n\n\
🕵️‍♀️**Spy**🕵️‍♀️\n\n\
A highly skilled infiltrator, the Spy has the ability to flip the votes during the approval phase of an expedition. \
She may only do this once, however, as everyone will be alerted that the votes have been flipped.\n\n\
Tip: The Spy needs to be strategic about when to flip the votes to favor the Warriors, as it can easily backfire if people do not vote the way you expect.'
            return tutorial9

        elif num == 10:
            tutorial10 = '**In-Game Effects**\n\n\
🔮**Ymir\'s Blessing**🔮\n\n\
Ymir\'s Blessing allows a player who has it to check another player\'s true allegiance (Warriors or Soldiers). At the start of the 3rd expedition, a random player will be given Ymir\'s blessing. \
They may use it on another player to find out their allegiance (will be DMed to them in private).\n\n\
Ymir\'s blessing will then be passed onto the player who was investigated, who can use it at the start of the next expedition, and so on. \
However, anyone who has had Ymir\'s blessing previously is granted immunity to being investigated by future holders of it.\n\n\
📢**Paths**📢\n\n\
The Paths ability allows players to make an announcement anonymously to everyone in the game. When activated, a random person will be designated as the Paths holder at the start of every expedition. \
They will then be able to send a single message to the game channel without revealing their role or identity.'
            return tutorial10

        elif num == 11:
            tutorial11 = '💡**Tips and Strategies**💡\n\n\
💭 A common strategy used by the Warriors is to succeed the first couple of expeditions. This allows them to maintain their cover for a longer period, but also \
gives the Soldiers a significant headstart. Be tactful about when to do this (for example, if people are starting to suspect you).\n\n\
💭 Warriors can often find clues on who the Coordinate is by looking at past Commanders\' decisions and the voting rounds. As the Coordinate knows the Warriors\' identities, \
they rarely ever pick or vote incorrectly.\n\n\
💭 It is helpful to also do weird and crazy things occasionally and be known as the guy/gal who just does unexpected things sometimes. This might be arrogantly displaying how they \
are randomly voting for an expedition for fun, vehemently accusing someone of being a Warrior with no solid evidence, or being mute \
for a game. This kind of randomness will help cover up slip-ups in games because it can be shrugged off as, "but he/she just always does that kind of thing".\n\n\
💭 Remember that at the end of the day, this is just a game! Lying and deception are part-and-parcel of the game; don\'t take it too seriously, and you\'ll have a great time!'
            return tutorial11
        




