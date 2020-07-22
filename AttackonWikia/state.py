import discord
from AttackonWikia import fetchURL
import time
import random
import math
from openpyxl import load_workbook

class State():
    def __init__(self):
        # Variables
        self.question_no = 0
        self.question_set = None
        self.clue_no = 0
        self.game_channel = None
        self.image = None
        self.challenge = False
        self.hangman_challenge = False
        self.players = []
        self.scores = [0,0]
        self.timer = 0
        self.correct = False
        self.letters_guessed = ''
        self.wrong_answers = 0
        self.wrong_letter = False

        # Constants
        self.intro_msg = discord.Embed(title = 'Game Reset!', description = 'Type **`~new`** to start a new puzzle.\nType **`~help`** to see the list of available commands.', colour = 0xC0C0C0)

        self.levelling_system = ([1,0],[2,1],[3,2],[4,3],[5,5],[6,7],[7,10],[8,13],[9,16],[10,20],[11,24],[12,28],[13,32],[14,36],[15,40],[16,45],[17,50],[18,55],[19,60],[20,66],[21,72],[22,78],[23,85],[24,92],[25,100],
        [26,108],[27,116],[28,124],[29,132],[30,141],[31,150],[32,160],[33,170],[34,180],[35,190],[36,200],[37,215],[38,230],[39,245],[40,260],[41,280],[42,300],[43,320],[44,350],[45,380],[46,410],[47,440],
        [48,480],[49,520],[50,560],[51,600],[52,650],[53,700],[54,750],[55,800],[56,875],[57,950],[58,1025],[59,1100],[60,1200],[61,1300],[62,1400],[63,1500],[64,1625],[65,1750],[66,1875],[67,2000],
        [68,2150],[69,2300],[70,2500],[71,2725],[72,2950],[73,3175],[74,3400],[75,3650],[76,3900],[77,4150],[78,4425],[79,4700],[80,5000],[81,5450],[82,5900],[83,6350],[84,6800],[85,7300],[86,7800],
        [87,8300],[88,8850],[89,9400],[90,10000],[91,10900],[92,11800],[93,12700],[94,13600],[95,14600],[96,15600],[97,16600],[98,17700],[99,18800],[100,20000])

        self.one_clue_achievements = {10:'Bronze', 25:'Silver', 80:'Gold', 150:'Platinum', 400:'Diamond', 750:'Master', 1500:'Grandmaster'}
        self.consecutive_days_achievements = {5:'Bronze', 14:'Silver', 30:'Gold', 60:'Platinum', 100:'Diamond', 200:'Master', 365:'Grandmaster'}
        self.challenge_achievements = {5:'Bronze', 15:'Silver', 30:'Gold', 50:'Platinum', 80:'Diamond', 150:'Master', 300:'Grandmaster'}
        self.hangman_achievements = {5:'Bronze', 15:'Silver', 50:'Gold', 100:'Platinum', 250:'Diamond', 500:'Master', 1000:'Grandmaster'}

        self.achievement_rewards = {'Bronze': 20, 'Silver': 50, 'Gold': 100, 'Platinum': 150, 'Diamond': 300, 'Master':500, 'Grandmaster':1000}
        self.badge_emojis = {'Bronze': '🥉', 'Silver': '🥈', 'Gold': '🥇', 'Platinum': '💠', 'Diamond': '💎', 'Master': '👑', 'Grandmaster':'🎓'}

    # Methods
    def game_reset(self):
        self.question_no = 0
        self.question_set = None
        self.clue_no = 0
        self.game_channel = None
        self.image = None
        self.challenge = False
        self.hangman_challenge = False
        self.players = []
        self.scores = [0,0]
        self.timer = 0
        self.correct = False
        self.letters_guessed = ''
        self.wrong_answers = 0
        self.wrong_letter = False

    def get_new_question(self):
        self.game_reset()
        new_qn_msgs = []
        self.question_set = fetchURL.new_question()
        self.clue_no = 1
        # +1 to record of questions asked
        self.new_question()
        new_qn_msgs.append('**Clue 1:**')
        new_qn_msgs.append(self.question_set['clues'][0])
        new_qn_msgs.append('Type `~clue` for the next hint.')

        return new_qn_msgs

    def get_clue(self):
        clue_msgs = []
        if self.question_set == None:
            clue_msgs.append('No puzzle is currently active!')
        elif len(self.question_set) == 2:
            clue_msgs.append('You can\'t use that in a hangman game!')
        elif 'clues' not in self.question_set:
            clue_msgs.append('You can\'t use that when guessing the image!')
        elif self.clue_no >= 5:
            clue_msgs.append('You have used up all the available clues!')
        else:
            self.clue_no += 1
            clue_msgs.append('**Clue ' + str(self.clue_no) + '**:')
            clue_msgs.append(self.question_set['clues'][self.clue_no - 1])

        return clue_msgs

    def get_new_image(self):
        self.game_reset()
        self.question_set = fetchURL.new_image()
        self.new_image_update()
        image_embed = discord.Embed(title = 'Guess the image!', colour = 0xC0C0C0)
        image_embed.set_image(url = self.question_set['image'])
        return image_embed

    def get_new_hangman(self):
        self.game_reset()
        self.question_set = fetchURL.new_hangman()
        # +1 to record of questions asked
        self.new_hangman_question()
        current_word = self.get_hangman_word()
        hangman_status = self.get_hangman_status()
        hangman_embed = discord.Embed(title = 'Hangman Mode', description = current_word + '\n\n' + hangman_status, colour = 0xC0C0C0)
        return hangman_embed

    def get_new_hangman_challenge(self):
        self.letters_guessed = ''
        current_word = self.get_hangman_word()
        hangman_embed = discord.Embed(title = 'Hangman Mode', description = current_word, colour = 0xC0C0C0)
        return hangman_embed

    def get_hangman_challenge_word(self):
        current_word = self.get_hangman_word()
        hangman_embed = discord.Embed(title = 'Hangman Mode', description = current_word, colour = 0xC0C0C0)
        return hangman_embed

    def hangman_reveal_letter(self):
        # Updates letters guessed with a randomly chosen letter
        valid_letters = list(filter(lambda x:x not in self.letters_guessed, self.question_set['title'].upper()))
        def remove_duplicates(word):
            all_letters = ''
            for i in word:
                if i not in all_letters:
                    all_letters += i
            return all_letters
        valid_letters = remove_duplicates(valid_letters)
        chosen_letter = valid_letters[random.randint(0, len(valid_letters)-1)]
        self.letters_guessed += chosen_letter + ' '
        
    def letter_guess_check(self, msg):
        if len(msg) != 1 or self.question_set == None:
            return False
        elif len(self.question_set) != 2:
            return False
        else:
            valid_letters = 'abcdefghijklmnopqrstuvwxyz'
            if msg.lower() in valid_letters:
                return True
            else:
                return False

    def letter_guess(self, letter, player):
        # Letter already used
        if letter.upper() in self.letters_guessed:
            self.wrong_letter = True

            description = '🔄 Letter `' + letter.upper() + '` has already been used. You can guess again in 1 second.'
            current_word = self.get_hangman_word()
            hangman_status = self.get_hangman_status()
            hangman_embed = discord.Embed(title = 'Hangman Mode', description = description + '\n' + current_word + '\n\n' + hangman_status, colour = 0xC0C0C0)
            hangman_embed.set_footer(text = self.letters_guessed)
            return hangman_embed
        
        # Correct letter
        elif letter.lower() in self.question_set['title'].lower():
            self.letters_guessed += letter.upper() + ' '
            # Word not completed
            if '◯' in self.get_hangman_word():
                description = '✅ ' + player.mention + ' has guessed a letter `' + letter.upper() + '`!'
                current_word = self.get_hangman_word()
                hangman_status = self.get_hangman_status()
                hangman_embed = discord.Embed(title = 'Hangman Mode', description = description + '\n' + current_word + '\n\n' + hangman_status, colour = 0xC0C0C0)
                hangman_embed.set_footer(text = self.letters_guessed)
                return hangman_embed

            # Word completed
            else:
                return self.correct_answer(player)

        # Wrong letter
        else:
            self.letters_guessed += letter.upper() + ' '
            self.wrong_letter = True
            self.wrong_answers += 1

            if self.wrong_answers < 6:
                description = '❌ Letter `' + letter.upper() + '` does not exist. You can guess again in 1 second.'
            else:
                description = '❌ Letter `' + letter.upper() + '` does not exist.'
            current_word = self.get_hangman_word()
            hangman_status = self.get_hangman_status()
            hangman_embed = discord.Embed(title = 'Hangman Mode', description = description + '\n' + current_word + '\n\n' + hangman_status, colour = 0xC0C0C0)
            hangman_embed.set_footer(text = self.letters_guessed)

            return hangman_embed

    def get_hangman_word(self):
        title = '`'
        for letter in self.question_set['title'].upper():
            if letter == ' ':
                title += ' \u2001 '
            elif letter in self.letters_guessed:
                title += letter + ' '
            else:
                title += '◯ '
 
        title += '`'
        return title

    def get_hangman_status(self):
        hangman_dict = {
            0: '| \u2004 🇲 \u2001\u2000 🇷 \u2001\u2000 🇸\n|🏡🏡 🏠🏠 🏰🏰',
            1: '| \u2004 🇲 \u2001\u2000 🇷 \u2001\u2000 🇸\n|💥🏡 🏠🏠 🏰🏰',
            2: '| \u2004 🇲 \u2001\u2000 🇷 \u2001\u2000 🇸\n|💥💥 🏠🏠 🏰🏰',
            3: '| \u2004 🇲 \u2001\u2000 🇷 \u2001\u2000 🇸\n|💥💥 💥🏠 🏰🏰',
            4: '| \u2004 🇲 \u2001\u2000 🇷 \u2001\u2000 🇸\n|💥💥 💥💥 🏰🏰',
            5: '| \u2004 🇲 \u2001\u2000 🇷 \u2001\u2000 🇸\n|💥💥 💥💥 💥🏰',
            6: '| \u2004 🇲 \u2001\u2000 🇷 \u2001\u2000 🇸\n|💥💥 💥💥 💥💥'
        }
        return hangman_dict[self.wrong_answers]

    def get_answer(self):
        answer_msgs = []
        if self.question_set == None:
            answer_msgs.append('No puzzle is currently active!')
        else:
            if self.wrong_answers >= 6:
                answer_msgs.append('Hangman! The page is **' + self.question_set['title'] + ' !**')
            else:
                answer_msgs.append('The page is **' + self.question_set['title'] + ' !**')
            answer_msgs.append(self.question_set['url'])
            answer_msgs.append('Type `~new` to start a new puzzle. Type `~answer` to reveal the solution. Type `~help` to see the list of available commands.')
            self.game_reset()

        return answer_msgs

    def correct_answer(self, player):
        correct_answer_msgs = []
        correct_answer_msgs.append('Correct! The page is **' + self.question_set['title'] + ' !**')
        correct_answer_msgs.append(self.question_set['url'])
        correct_answer_msgs.append('Type `~new` to start a new puzzle. Type `~answer` to reveal the solution. Type `~help` to see the list of available commands.')
        
        # Log result in achievements and sends message if obtained achievement
        achievements = self.log_achievements(player)
        for achievement in achievements:
            correct_answer_msgs.append(achievement)
        
        # Update player records
        level_update = self.update_player_records(player)
        if level_update != None:
            correct_answer_msgs.append(level_update)

        # Update dailies
        dailies = self.update_dailies(player)
        for daily in dailies:
            correct_answer_msgs.append(daily)

        # +1 to global record of correct questions
        if len(self.question_set) == 3:
            if 'image' in self.question_set:
                self.correct_image()
            else:
                self.correct_question()
        elif len(self.question_set) == 2:
            self.correct_hangman_question()

        self.game_reset()

        return correct_answer_msgs

    def new_challenge(self, msg_channel):
        self.challenge = True
        if self.hangman_challenge == False:
            self.question_no = 1
            self.question_set = fetchURL.new_question()
            self.clue_no = 1
            self.game_channel = msg_channel
        else:
            # clue_no stays at 0
            self.question_no = 1
            self.question_set = fetchURL.new_hangman()
            self.game_channel = msg_channel

    def new_question(self):
        wb = load_workbook("AttackonWikia/records.xlsx")
        game_stats = wb['Overall']
        game_stats['A2'] = game_stats['A2'].value + 1 
        wb.save("AttackonWikia/records.xlsx")

    def new_image_update(self):
        wb = load_workbook("AttackonWikia/records.xlsx")
        game_stats = wb['Overall']
        game_stats['H2'] = game_stats['H2'].value + 1 
        wb.save("AttackonWikia/records.xlsx")

    def new_hangman_question(self):
        wb = load_workbook("AttackonWikia/records.xlsx")
        game_stats = wb['Overall']
        game_stats['F2'] = game_stats['F2'].value + 1 
        wb.save("AttackonWikia/records.xlsx")

    def correct_question(self):
        wb = load_workbook("AttackonWikia/records.xlsx")
        game_stats = wb['Overall']
        game_stats['B2'] = game_stats['B2'].value + 1 
        wb.save("AttackonWikia/records.xlsx")
    
    def correct_image(self):
        wb = load_workbook("AttackonWikia/records.xlsx")
        game_stats = wb['Overall']
        game_stats['I2'] = game_stats['I2'].value + 1
        wb.save("AttackonWikia/records.xlsx")

    def correct_hangman_question(self):
        wb = load_workbook("AttackonWikia/records.xlsx")
        game_stats = wb['Overall']
        game_stats['G2'] = game_stats['G2'].value + 1 
        wb.save("AttackonWikia/records.xlsx")

    def new_challenge_qn(self):
        wb = load_workbook("AttackonWikia/records.xlsx")
        game_stats = wb['Overall']
        if self.hangman_challenge == False:
            game_stats['A2'] = game_stats['A2'].value + 1 
        else:
            game_stats['F2'] = game_stats['F2'].value + 1 
        game_stats['D2'] = game_stats['D2'].value + 1 
        wb.save("AttackonWikia/records.xlsx")

    def correct_challenge_qn(self):
        wb = load_workbook("AttackonWikia/records.xlsx")
        game_stats = wb['Overall']
        if self.hangman_challenge == False:
            game_stats['B2'] = game_stats['B2'].value + 1 
        else:
            game_stats['G2'] = game_stats['G2'].value + 1 
        game_stats['E2'] = game_stats['E2'].value + 1 
        wb.save("AttackonWikia/records.xlsx")

    def finish_challenge(self):
        wb = load_workbook("AttackonWikia/records.xlsx")
        game_stats = wb['Overall']
        game_stats['C2'] = game_stats['C2'].value + 1
        wb.save("AttackonWikia/records.xlsx")

    def update_player_records(self, user_obj):
        wb = load_workbook("AttackonWikia/records.xlsx")
        player_stats = wb['Players']
        inside = False
        i = 0
        for row in player_stats:
            i += 1
            # Player exists in the records
            if row[0].value == str(user_obj.id):
                inside = True
                # Old player level
                for level in self.levelling_system:
                    if player_stats['B' + str(i)].value < level[1]:
                        break
                    old_player_level = level[0]

                # +1 to exp and correct answers
                player_stats['B' + str(i)] = player_stats['B' + str(i)].value + 1
                if len(self.question_set) == 3:
                    if 'image' in self.question_set:
                        player_stats['G' + str(i)] = player_stats['G' + str(i)].value + 1
                    else:
                        player_stats['C' + str(i)] = player_stats['C' + str(i)].value + 1
                elif len(self.question_set) == 2:
                    player_stats['F' + str(i)] = player_stats['F' + str(i)].value + 1

                # New player level
                for level in self.levelling_system:
                    if player_stats['B' + str(i)].value < level[1]:
                        break
                    new_player_level = level[0]

                wb.save("AttackonWikia/records.xlsx")

                if new_player_level > old_player_level:
                    return '🆙 | Congratulations ' + user_obj.mention + '! You have reached level **' + str(new_player_level) + '**!'
                else:
                    return None

        if inside == False:
            if len(self.question_set) == 3:
                if 'image' in self.question_set:
                    player_stats.append([str(user_obj.id), 1, 0, 0, 0, 0, 1])
                else:
                    player_stats.append([str(user_obj.id), 1, 1, 0, 0, 0, 0])
            elif len(self.question_set) == 2:
                player_stats.append([str(user_obj.id), 1, 0, 0, 0, 1, 0])
            wb.save("AttackonWikia/records.xlsx")
            return '🆙 | Congratulations ' + user_obj.mention + '! You have reached level **2**!'

    def get_score(self, num):
        score = self.scores[num]
        if score == 0:
            return '⬛ ⬛ ⬛'
        elif score == 1:
            return '✅ ⬛ ⬛'
        elif score == 2:
            return '✅ ✅ ⬛'
        elif score == 3:
            return '✅ ✅ ✅'

    def get_scoreboard(self):
        scoreboard = discord.Embed(title='Scores\n', colour=0xB4B4B4)
        scoreboard.add_field(name = self.players[0].name, value = self.get_score(0)) 
        scoreboard.add_field(name = self.players[1].name, value = self.get_score(1)) 
        return scoreboard

    def next_question(self):
        if self.hangman_challenge == False:
            self.question_set = fetchURL.new_question()
            self.clue_no = 1
        else:
            self.question_set = fetchURL.new_hangman()
        self.timer = 0
        self.correct = False

    def get_winner(self):
        winner_index = 0 if self.scores[0] >= 3 else 1
        loser_index = 1 if self.scores[0] >= 3 else 0
        players = [self.players[winner_index], self.players[loser_index]]

        scoreboard = discord.Embed(title='🏆 The winner is **' + players[0].name + '**! 🏆', colour=0xB4B4B4)
        scoreboard.set_thumbnail(url = players[0].avatar_url)
        scoreboard.add_field(name = self.players[0].name, value = self.get_score(0)) 
        scoreboard.add_field(name = self.players[1].name, value = self.get_score(1)) 
        return scoreboard

    def challenge_update_player_records(self):
        winner_index = 0 if self.scores[0] >= 3 else 1
        loser_index = 1 if self.scores[0] >= 3 else 0
        players = [self.players[winner_index], self.players[loser_index]]
        
        wb = load_workbook("AttackonWikia/records.xlsx")
        player_stats = wb['Players']

        level_updates = []

        j = 0
        for player in players:
            j += 1
            i = 0
            inside = False
            for row in player_stats:
                i += 1
                if row[0].value == str(player.id):
                    inside = True
                    # Old player level
                    for level in self.levelling_system:
                        if player_stats['B' + str(i)].value < level[1]:
                            break
                        old_player_level = level[0]
        
                    # Add to exp and challenge won/completed
                    if j == 1:
                        player_stats['B' + str(i)] = player_stats['B' + str(i)].value + 10
                        player_stats['E' + str(i)] = player_stats['E' + str(i)].value + 1
                        player_stats['D' + str(i)] = player_stats['D' + str(i)].value + 1
                    elif j == 2:
                        player_stats['B' + str(i)] = player_stats['B' + str(i)].value + 5
                        player_stats['D' + str(i)] = player_stats['D' + str(i)].value + 1

                    # New player level
                    for level in self.levelling_system:
                        if player_stats['B' + str(i)].value < level[1]:
                            break
                        new_player_level = level[0]

                    wb.save("AttackonWikia/records.xlsx")

                    if new_player_level > old_player_level:
                        level_updates.append('🆙 | Congratulations ' + player.mention + '! You have reached level **' + str(new_player_level) + '**!') 

                    break

            if inside == False:
                player_stats.append([str(player.id), 5, 0, 0, 0, 0])
                level_updates.append('🆙 | Congratulations ' + player.mention + '! You have reached level **5**!')
                wb.save("AttackonWikia/records.xlsx")

        return level_updates

    def update_dailies(self, *player):
        wb = load_workbook("AttackonWikia/records.xlsx")
        player_dailies = wb['Dailies']
        player_stats = wb['Players']
        cur_date = str(time.gmtime().tm_year) + '-' + str(time.gmtime().tm_mon) + '-' + str(time.gmtime().tm_mday)
        cur_dailies = []

        if player:
            i = 0
            inside = False
            for row in player_dailies:
                i += 1
                if row[0].value == str(player[0].id):
                    inside = True
                    # B: prev play, C: Standard, D: Hangman, E: Challenge, F: Image
                    prev_date = player_dailies['B' + str(i)].value
                    # Still same day
                    if cur_date == prev_date:
                        if len(self.question_set) == 3:
                            # Image question
                            if 'image' in self.question_set:
                                player_dailies['F' + str(i)] = player_dailies['F' + str(i)].value + 1
                            # Normal question
                            else:
                                player_dailies['C' + str(i)] = player_dailies['C' + str(i)].value + 1
        
                        # Hangman
                        else:
                            player_dailies['D' + str(i)] = player_dailies['D' + str(i)].value + 1
                            
                        wb.save("AttackonWikia/records.xlsx")

                        # Award daily if reached 10 points
                        if (len(self.question_set) == 3 and 'image' not in self.question_set and player_dailies['C' + str(i)].value == 10) or \
                        (len(self.question_set) == 2 and player_dailies['D' + str(i)].value == 10) or \
                        (len(self.question_set) == 3 and 'image' in self.question_set and player_dailies['F' + str(i)].value == 10):
                            # Daily reward msg
                            if len(self.question_set) == 3:
                                if 'image' in self.question_set:
                                    daily_reward_msg = '☑ | Congratulations, ' + player[0].mention + ', you have completed the Images daily task! **(+15 Exp)**'
                                else:
                                    daily_reward_msg = '☑ | Congratulations, ' + player[0].mention + ', you have completed the standard games daily task! **(+15 Exp)**'
                            else:
                                daily_reward_msg = '☑ | Congratulations, ' + player[0].mention + ', you have completed the Hangman daily task! **(+15 Exp)**'
                            cur_dailies.append(daily_reward_msg)

                            # Add exp to player
                            j = 0
                            for row in player_stats:
                                j += 1
                                if row[0].value == str(player[0].id):
                                    # Old player level
                                    for level in self.levelling_system:
                                        if player_stats['B' + str(j)].value < level[1]:
                                            break
                                        old_player_level = level[0]

                                    # Add exp
                                    player_stats['B' + str(j)] = player_stats['B' + str(j)].value + 15

                                    # New player level
                                    for level in self.levelling_system:
                                        if player_stats['B' + str(j)].value < level[1]:
                                            break
                                        new_player_level = level[0]

                                    wb.save("AttackonWikia/records.xlsx")

                                    if new_player_level > old_player_level:
                                        cur_dailies.append('🆙 | Congratulations ' + player[0].mention + '! You have reached level **' + str(new_player_level) + '**!')

                                    break                    
                                    
                    # New day, reset
                    else:
                        if len(self.question_set) == 3:
                            if 'image' in self.question_set:
                                player_dailies['C' + str(i)] = 0
                                player_dailies['D' + str(i)] = 0
                                player_dailies['F' + str(i)] = 1
                            else:
                                player_dailies['C' + str(i)] = 1
                                player_dailies['D' + str(i)] = 0
                                player_dailies['F' + str(i)] = 0
                        else:
                            player_dailies['C' + str(i)] = 0
                            player_dailies['D' + str(i)] = 1
                            player_dailies['F' + str(i)] = 0
                        player_dailies['B' + str(i)] = cur_date
                        player_dailies['E' + str(i)] = 0
                        wb.save("AttackonWikia/records.xlsx")

            if inside == False:
                if len(self.question_set) == 3:
                    # Image
                    if 'image' in self.question_set:
                        player_dailies.append([str(player[0].id), cur_date, 0, 0, 0, 1])
                    # Normal question
                    else:
                        player_dailies.append([str(player[0].id), cur_date, 1, 0, 0, 0])
                # Hangman
                else:
                    player_dailies.append([str(player[0].id), cur_date, 0, 1, 0, 0])
                wb.save("AttackonWikia/records.xlsx")
                            
        # Challenge completed, both players
        else:
            for player in self.players:
                i = 0
                inside = False
                for row in player_dailies:
                    i += 1
                    if row[0].value == str(player.id):
                        inside = True
                        # B: prev play, C: Standard, D: Hangman, E: Challenge
                        prev_date = player_dailies['B' + str(i)].value
                        # Still same day
                        if cur_date == prev_date:
                            player_dailies['E' + str(i)] = player_dailies['E' + str(i)].value + 1
                            wb.save("AttackonWikia/records.xlsx")

                        # New day, reset
                        else:
                            player_dailies['B' + str(i)] = cur_date
                            player_dailies['C' + str(i)] = 0
                            player_dailies['D' + str(i)] = 0
                            player_dailies['E' + str(i)] = 1
                            player_dailies['F' + str(i)] = 0
                            wb.save("AttackonWikia/records.xlsx")

                        # Award daily, by default
                        if player_dailies['E' + str(i)].value == 1:
                            daily_reward_msg = '☑ | Congratulations, ' + player.mention + ', you have completed the Challenge daily task! **(+25 Exp)**'
                            cur_dailies.append(daily_reward_msg)

                            # Add exp to player
                            j = 0
                            for row in player_stats:
                                j += 1
                                if row[0].value == str(player.id):
                                    # Old player level
                                    for level in self.levelling_system:
                                        if player_stats['B' + str(j)].value < level[1]:
                                            break
                                        old_player_level = level[0]

                                    # Add exp
                                    player_stats['B' + str(j)] = player_stats['B' + str(j)].value + 25

                                    # New player level
                                    for level in self.levelling_system:
                                        if player_stats['B' + str(j)].value < level[1]:
                                            break
                                        new_player_level = level[0]

                                    wb.save("AttackonWikia/records.xlsx")

                                    if new_player_level > old_player_level:
                                        cur_dailies.append('🆙 | Congratulations ' + player.mention + '! You have reached level **' + str(new_player_level) + '**!')

                                    break

                if inside == False:
                    player_dailies.append([str(player.id), cur_date, 0, 0, 1, 0])
                    wb.save("AttackonWikia/records.xlsx")
                    # Award daily, by default
                    daily_reward_msg = '☑ | Congratulations, ' + player.mention + ', you have completed the Challenge daily task! **(+25 Exp)**'
                    cur_dailies.append(daily_reward_msg)

                    # Add exp to player
                    j = 0
                    for row in player_stats:
                        j += 1
                        if row[0].value == str(player.id):
                            # Old player level
                            for level in self.levelling_system:
                                if player_stats['B' + str(j)].value < level[1]:
                                    break
                                old_player_level = level[0]

                            # Add exp
                            player_stats['B' + str(j)] = player_stats['B' + str(j)].value + 25

                            # New player level
                            for level in self.levelling_system:
                                if player_stats['B' + str(j)].value < level[1]:
                                    break
                                new_player_level = level[0]

                            wb.save("AttackonWikia/records.xlsx")

                            if new_player_level > old_player_level:
                                cur_dailies.append('🆙 | Congratulations ' + player.mention + '! You have reached level **' + str(new_player_level) + '**!')

                            break
                
        return cur_dailies
    
    def next_day_check(self, last_date, cur_date):
        endday = {1:31, 2:28, 3:31, 4:30, 5:31, 6:30, 7:31, 8:31, 9:30, 10:31, 11:30, 12:31}
        last = list(map(lambda x: int(x), last_date.split('-')))
        cur = list(map(lambda x: int(x), cur_date.split('-')))
        if cur[2] - last[2] == 1:
            return True
        elif cur[1] - last[1] == 1 and (cur[2] == 1 and last[2] >= endday[last[1]]):
            return True
        elif cur[0] - last[0] == 1 and (cur[1] == 1 and cur[2] == 1) and (last[1] == 12 and last[2] == 31):
            return True
        else:
            return False

    def log_achievements(self, player):
        # Increment achievements progress for questions and returns achievement messages if any
        wb = load_workbook("AttackonWikia/records.xlsx")
        achievements = wb['Achievements']
        player_stats = wb['Players']

        achievement_msgs = []
        inside = False
        i = 0
        for row in achievements:
            i += 1
            # Player exists in the records
            if row[0].value == str(player.id):
                inside = True
                # 1 clue
                if self.clue_no == 1:
                    achievements['B' + str(i)] = achievements['B' + str(i)].value + 1

                    # If achievement earned, add exp and append achievement msg
                    if achievements['B' + str(i)].value in self.one_clue_achievements or (achievements['B' + str(i)].value > 1500 and achievements['H' + str(i)].value == 0):
                        achievement_value = achievements['B' + str(i)].value
                        if achievements['B' + str(i)].value >= 1500 and achievements['H' + str(i)].value == 0:
                            achievements['H' + str(i)] = 1
                            achievement_value = 1500

                        # Achievement msg
                        one_clue = self.badge_emojis[self.one_clue_achievements[achievement_value]] + ' | Congratulations ' + player.mention + \
                        '! You have obtained ' + ('a ' if 'Master' not in self.one_clue_achievements[achievement_value] else 'the ') + '**' + \
                        self.one_clue_achievements[achievement_value] + (' of Archives' if 'Master' in self.one_clue_achievements[achievement_value] else '') + \
                        '** badge for answering **' + str(achievement_value) + '** questions correctly with only 1 clue! **(+' + \
                        str(self.achievement_rewards[self.one_clue_achievements[achievement_value]]) + ' Exp)**'
                        
                        achievement_msgs.append(one_clue)

                        # Add exp to player
                        j = 0
                        for row in player_stats:
                            j += 1
                            if row[0].value == str(player.id):
                                # Old player level
                                for level in self.levelling_system:
                                    if player_stats['B' + str(j)].value < level[1]:
                                        break
                                    old_player_level = level[0]

                                # Add exp
                                player_stats['B' + str(j)] = player_stats['B' + str(j)].value + self.achievement_rewards[self.one_clue_achievements[achievement_value]]

                                # New player level
                                for level in self.levelling_system:
                                    if player_stats['B' + str(j)].value < level[1]:
                                        break
                                    new_player_level = level[0]

                                wb.save("AttackonWikia/records.xlsx")

                                if new_player_level > old_player_level:
                                    achievement_msgs.append('🆙 | Congratulations ' + player.mention + '! You have reached level **' + str(new_player_level) + '**!')

                                break

                # Perfect hangman
                if len(self.question_set) == 2 and self.wrong_answers == 0 and self.challenge == False:
                    achievements['G' + str(i)] = achievements['G' + str(i)].value + 1

                    # If achievement earned, add exp and append achievement msg
                    if achievements['G' + str(i)].value in self.hangman_achievements or (achievements['G' + str(i)].value > 1000 and achievements['K' + str(i)].value == 0):
                        achievement_value = achievements['G' + str(i)].value
                        if achievements['G' + str(i)].value >= 1000 and achievements['K' + str(i)].value == 0:
                            achievements['K' + str(i)] = 1
                            achievement_value = 1000

                        # Achievement msg
                        one_clue = self.badge_emojis[self.hangman_achievements[achievement_value]] + ' | Congratulations ' + player.mention + \
                        '! You have obtained ' + ('a ' if 'Master' not in self.hangman_achievements[achievement_value] else 'the ') + '**' + \
                        self.hangman_achievements[achievement_value] + (' of Hangman' if 'Master' in self.hangman_achievements[achievement_value] else '') + \
                        '** badge for winning **' + str(achievement_value) + '** hangman games without any mistakes! **(+' + \
                        str(self.achievement_rewards[self.hangman_achievements[achievement_value]]) + ' Exp)**'
                        
                        achievement_msgs.append(one_clue)

                        # Add exp to player
                        j = 0
                        for row in player_stats:
                            j += 1
                            if row[0].value == str(player.id):
                                # Old player level
                                for level in self.levelling_system:
                                    if player_stats['B' + str(j)].value < level[1]:
                                        break
                                    old_player_level = level[0]

                                # Add exp
                                player_stats['B' + str(j)] = player_stats['B' + str(j)].value + self.achievement_rewards[self.hangman_achievements[achievement_value]]

                                # New player level
                                for level in self.levelling_system:
                                    if player_stats['B' + str(j)].value < level[1]:
                                        break
                                    new_player_level = level[0]

                                wb.save("AttackonWikia/records.xlsx")

                                if new_player_level > old_player_level:
                                    achievement_msgs.append('🆙 | Congratulations ' + player.mention + '! You have reached level **' + str(new_player_level) + '**!')

                                break

                # Consecutive days
                cur_date = str(time.gmtime().tm_year) + '-' + str(time.gmtime().tm_mon) + '-' + str(time.gmtime().tm_mday)
                last_date = achievements['D' + str(i)].value
                if cur_date != last_date:
                    # Update newest date
                    achievements['D' + str(i)] = cur_date

                    # 1 day
                    if self.next_day_check(last_date, cur_date):
                        achievements['E' + str(i)] = achievements['E' + str(i)].value + 1
                        if achievements['E' + str(i)].value > achievements['F' + str(i)].value:
                            # New max streak
                            achievements['F' + str(i)] = achievements['E' + str(i)].value
                        
                            # If achievement earned, add exp and append achievement msg
                            if achievements['F' + str(i)].value in self.consecutive_days_achievements or (achievements['F' + str(i)].value > 365 and achievements['J' + str(i)].value == 0):
                                achievement_value = achievements['F' + str(i)].value
                                if achievements['F' + str(i)].value >= 365 and achievements['J' + str(i)].value == 0:
                                    achievements['J' + str(i)] = 1
                                    achievement_value = 365

                                # Achievement msg
                                consecutive = self.badge_emojis[self.consecutive_days_achievements[achievement_value]] + ' | Congratulations ' + player.mention + \
                                '! You have obtained a **' + self.consecutive_days_achievements[achievement_value] + '** badge for playing the game every day for **' + \
                                str(achievement_value) + '** days! **(+' + str(self.achievement_rewards[self.consecutive_days_achievements[achievement_value]]) + ' Exp)**'
                                
                                achievement_msgs.append(consecutive)

                                # Add exp to player
                                j = 0
                                for row in player_stats:
                                    j += 1
                                    if row[0].value == str(player.id):
                                        # Old player level
                                        for level in self.levelling_system:
                                            if player_stats['B' + str(j)].value < level[1]:
                                                break
                                            old_player_level = level[0]

                                        # Add exp
                                        player_stats['B' + str(j)] = player_stats['B' + str(j)].value + self.achievement_rewards[self.consecutive_days_achievements[achievement_value]]

                                        # New player level
                                        for level in self.levelling_system:
                                            if player_stats['B' + str(j)].value < level[1]:
                                                break
                                            new_player_level = level[0]

                                        wb.save("AttackonWikia/records.xlsx")

                                        if new_player_level > old_player_level:
                                            achievement_msgs.append('🆙 | Congratulations ' + player.mention + '! You have reached level **' + str(new_player_level) + '**!')

                                        break

                    # > 1 day
                    else:
                        achievements['E' + str(i)] = 1
                    
                wb.save("AttackonWikia/records.xlsx")
                break

        if inside == False:
            cur_date = str(time.gmtime().tm_year) + '-' + str(time.gmtime().tm_mon) + '-' + str(time.gmtime().tm_mday)
            achievements.append([str(player.id), 0, 0, cur_date, 1, 1, 0])
            wb.save("AttackonWikia/records.xlsx")

        return achievement_msgs

    def log_challenge_achievements(self):
        winner_index = 0 if self.scores[0] >= 3 else 1
        loser_index = 1 if self.scores[0] >= 3 else 0
        players = [self.players[winner_index], self.players[loser_index]]

        wb = load_workbook("AttackonWikia/records.xlsx")
        achievements = wb['Achievements']
        player_stats = wb['Players']

        achievement_msgs = []

        for player in players:
            i = 0
            inside = False
            for row in achievements:
                i += 1
                # Player exists in the records
                if row[0].value == str(player.id):
                    inside = True
                    # Challenge played
                    achievements['C' + str(i)] = achievements['C' + str(i)].value + 1

                    # If achievement earned, add exp and append achievement msg
                    if achievements['C' + str(i)].value in self.challenge_achievements or (achievements['C' + str(i)].value > 300 and achievements['I' + str(i)].value == 0):
                        achievement_value = achievements['C' + str(i)].value
                        if achievements['C' + str(i)].value >= 300 and achievements['I' + str(i)].value == 0:
                            achievements['I' + str(i)] = 1
                            achievement_value = 300

                        # Achievement msg
                        challenge = self.badge_emojis[self.challenge_achievements[achievement_value]] + ' | Congratulations ' + player.mention + \
                        '! You have obtained ' + ('a ' if 'Master' not in self.challenge_achievements[achievement_value] else 'the ') + '**' + \
                        self.challenge_achievements[achievement_value] + (' of Challenges' if 'Master' in self.challenge_achievements[achievement_value] else '') + \
                        '** badge for having played **' + str(achievement_value) + \
                        '** challenges! **(+' + str(self.achievement_rewards[self.challenge_achievements[achievement_value]]) + ' Exp)**'
                        
                        achievement_msgs.append(challenge)

                        # Add exp to player
                        j = 0
                        for row in player_stats:
                            j += 1
                            if row[0].value == str(player.id):
                                # Old player level
                                for level in self.levelling_system:
                                    if player_stats['B' + str(j)].value < level[1]:
                                        break
                                    old_player_level = level[0]

                                # Add exp
                                player_stats['B' + str(j)] = player_stats['B' + str(j)].value + self.achievement_rewards[self.challenge_achievements[achievement_value]]

                                # New player level
                                for level in self.levelling_system:
                                    if player_stats['B' + str(j)].value < level[1]:
                                        break
                                    new_player_level = level[0]

                                wb.save("AttackonWikia/records.xlsx")

                                if new_player_level > old_player_level:
                                    achievement_msgs.append('🆙 | Congratulations ' + player.mention + '! You have reached level **' + str(new_player_level) + '**!')

                                break
                        
                    wb.save("AttackonWikia/records.xlsx")
                    break

            if inside == False:
                cur_date = str(time.gmtime().tm_year) + '-' + str(time.gmtime().tm_mon) + '-' + str(time.gmtime().tm_mday)
                achievements.append([str(player.id), 0, 1, cur_date, 1, 1, 0])
                wb.save("AttackonWikia/records.xlsx")

        return achievement_msgs

    def get_profile(self, player, server):
        # Returns profile of player in an embed
        wb = load_workbook("AttackonWikia/records.xlsx")
        player_data = wb['Players']

        # Extract data
        inside = False
        for row in player_data:
            if row[0].value == str(player.id):
                player_info = [player.name, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value]
                inside = True
                break
        if inside == False:
            player_info = [player.name, 0, 0, 0, 0, 0, 0]
            player_data.append([str(player.id), 0, 0, 0, 0, 0, 0])
        
        wb.save("AttackonWikia/records.xlsx")

        # Put info into embed
        for level in self.levelling_system:
            if player_info[1] < level[1]:
                next_exp = level[1] - player_info[1]
                break
            player_level = level[0]

        player_desc = 'Level: **' + str(player_level) + '**\nExperience Points: **' + str(player_info[1]) + '**'
        if len(self.levelling_system) > player_level >= 10:
            player_desc += ' (' + str(next_exp) + ' Exp to next level)'

        # Get total players in server
        player_rankings = [] # [[player id, exp], [...], ...]
        # Put all players into a list
        for row in player_data:
            if row[0].value == 'ID':
                continue
            else:
                player_rankings.append([row[0].value, row[1].value])
        
        # Sort by Exp from biggest to smallest
        player_rankings.sort(key = lambda x: x[1], reverse = True)

        # Get players in server
        server_users = list(map(lambda y: str(y.id), server.members))
        server_players = list(filter(lambda x:x[0] in server_users, player_rankings))
        total_players = len(server_players)

        # Get ranking
        i = 1
        for person in server_players:
            if person[0] == str(player.id):
                rank = i
                break
            else:
                i += 1

        player_profile = discord.Embed(title = 'Attack on Wikia Profile', description = player_desc, colour = 0xC0C0C0)
        player_profile.set_author(name = player_info[0], icon_url = player.avatar_url)
        player_profile.set_thumbnail(url = player.avatar_url)

        player_profile.add_field(name = 'Rank', value = str(rank) + '/' + str(total_players), inline = False)
        player_profile.add_field(name = 'Questions answered correctly', value = str(player_info[2]), inline = False)
        player_profile.add_field(name = 'Hangman games won', value = str(player_info[5]), inline = False)
        player_profile.add_field(name = 'Images guessed correctly', value = str(player_info[6]), inline = False)
        player_profile.add_field(name = 'Challenges played', value = str(player_info[3]))
        player_profile.add_field(name = 'Challenges won', value = str(player_info[4]))
        

        return player_profile

    def get_achievements(self, player):
        # Returns profile of player in an embed
        wb = load_workbook("AttackonWikia/records.xlsx")
        achievements = wb['Achievements']

        # Extract data
        inside = False
        for row in achievements:
            if row[0].value == str(player.id):
                badges = [player.name, row[1].value, row[2].value, row[4].value, row[5].value, row[6].value]
                inside = True
                break
        if inside == False:
            cur_date = str(time.gmtime().tm_year) + '-' + str(time.gmtime().tm_mon) + '-' + str(time.gmtime().tm_mday)
            badges = [player.name, 0, 0, 1, 1, 0]
            achievements.append([str(player.id), 0, 0, cur_date, 1, 1, 0, 0, 0, 0, 0])

        wb.save("AttackonWikia/records.xlsx")

        # Put info into embed
        num_badges = 0

        cur_days = badges[3]
        max_days = badges[4]
        consecutive_badges = ''
        consecutive_next = str(max_days)
        for step in sorted(list(self.consecutive_days_achievements)):
            if max_days < step:
                consecutive_next = str(cur_days) + '/' + str(step)
                break
            consecutive_badges += self.badge_emojis[self.consecutive_days_achievements[step]]
            num_badges += 1

        one_clue = badges[1]
        one_clue_badges = ''
        one_clue_next = str(one_clue)
        for step in sorted(list(self.one_clue_achievements)):
            if one_clue < step:
                one_clue_next = str(one_clue) + '/' + str(step)
                break
            one_clue_badges += self.badge_emojis[self.one_clue_achievements[step]]
            num_badges += 1

        challenges_played = badges[2]
        challenge_badges = ''
        challenge_next = str(challenges_played)
        for step in sorted(list(self.challenge_achievements)):
            if challenges_played < step:
                challenge_next = str(challenges_played) + '/' + str(step)
                break
            challenge_badges += self.badge_emojis[self.challenge_achievements[step]]
            num_badges += 1

        perfect_hangman = badges[5]
        perfect_hangman_badges = ''
        perfect_hangman_next = str(perfect_hangman)
        for step in sorted(list(self.hangman_achievements)):
            if perfect_hangman < step:
                perfect_hangman_next = str(perfect_hangman) + '/' + str(step)
                break
            perfect_hangman_badges += self.badge_emojis[self.hangman_achievements[step]]
            num_badges += 1

        player_achievements = discord.Embed(title = 'Badges for Attack on Wikia', description = 'Number of badges collected: **' + str(num_badges) + '**', colour = 0xC0C0C0)
        player_achievements.set_author(name = player.name, icon_url = player.avatar_url)
        player_achievements.set_thumbnail(url = player.avatar_url)

        player_achievements.add_field(name = 'Number of consecutive days played', value = consecutive_next)
        player_achievements.add_field(name = '⠀', value='⠀')
        player_achievements.add_field(name = 'Badges', value = consecutive_badges if consecutive_badges != '' else '-')
        player_achievements.add_field(name = 'Correct answers with 1 clue', value = one_clue_next)
        player_achievements.add_field(name = '⠀', value='⠀')
        player_achievements.add_field(name = 'Badges', value = one_clue_badges if one_clue_badges != '' else '-')
        player_achievements.add_field(name = 'Perfect Hangman games', value = perfect_hangman_next)
        player_achievements.add_field(name = '⠀', value='⠀')
        player_achievements.add_field(name = 'Badges', value = perfect_hangman_badges if perfect_hangman_badges != '' else '-')
        player_achievements.add_field(name = 'Challenges played', value = challenge_next)
        player_achievements.add_field(name = '⠀', value='⠀')
        player_achievements.add_field(name = 'Badges', value = challenge_badges if challenge_badges != '' else '-')

        return player_achievements

    def get_dailies(self, player):
        wb = load_workbook("AttackonWikia/records.xlsx")
        player_dailies = wb['Dailies']
        cur_date = str(time.gmtime().tm_year) + '-' + str(time.gmtime().tm_mon) + '-' + str(time.gmtime().tm_mday)
        cur_dailies = []

        def get_time_remaining(cur_hr, cur_min):
            # 24 hrs - current time
            cur_mins = cur_hr * 60 + cur_min
            rem_min = 24 * 60 - cur_mins
            hr = str(rem_min // 60)
            minute = str(rem_min % 60)
            return '[' + hr + 'h ' + minute + 'm]'

        i = 0
        inside = False
        for row in player_dailies:
            i += 1
            if row[0].value == str(player.id):
                inside = True
                # B: prev play, C: Standard, D: Hangman, E: Challenge
                prev_date = player_dailies['B' + str(i)].value
                if cur_date != prev_date:
                    # Reset
                    player_dailies['B' + str(i)] = cur_date
                    player_dailies['C' + str(i)] = 0
                    player_dailies['D' + str(i)] = 0
                    player_dailies['E' + str(i)] = 0
                    player_dailies['F' + str(i)] = 0
                    wb.save("AttackonWikia/records.xlsx")

                time_remaining = get_time_remaining(time.gmtime().tm_hour, time.gmtime().tm_min)
                if player_dailies['C' + str(i)].value >= 10:
                    standard_daily = '☑ Complete 10 standard puzzles! **[+15 Exp]**'
                    standard_daily2 = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ - Progress: **' + str(player_dailies['C' + str(i)].value) + '/10**'
                else:
                    standard_daily = '⬛ Complete 10 standard puzzles! **[+15 Exp]**'
                    standard_daily2 = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ - Progress: **' + str(player_dailies['C' + str(i)].value) + '/10**'
                    
                if player_dailies['D' + str(i)].value >= 10:
                    hangman_daily = '☑ Win 10 Hangman games! **[+15 Exp]**'
                    hangman_daily2 = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ - Progress: **' + str(player_dailies['D' + str(i)].value) + '/10**'
                else:
                    hangman_daily = '⬛ Win 10 Hangman games! **[+15 Exp]**'
                    hangman_daily2 = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ - Progress: **' + str(player_dailies['D' + str(i)].value) + '/10**'

                if player_dailies['F' + str(i)].value >= 10:
                    image_daily = '☑ Guess 10 images correctly! **[+15 Exp]**'
                    image_daily2 = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ - Progress: **' + str(player_dailies['F' + str(i)].value) + '/10**'
                else:
                    image_daily = '⬛ Guess 10 images correctly! **[+15 Exp]**'
                    image_daily2 = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ - Progress: **' + str(player_dailies['F' + str(i)].value) + '/10**'

                if player_dailies['E' + str(i)].value >= 1:
                    challenge_daily = '☑ Complete a challenge! **[+25 Exp]**'
                    challenge_daily2 = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ - Progress: **' + str(player_dailies['E' + str(i)].value) + '/1**'
                else:
                    challenge_daily = '⬛ Complete a challenge! **[+25 Exp]**'
                    challenge_daily2 = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ - Progress: **' + str(player_dailies['E' + str(i)].value) + '/1**'

        if inside == False:
            player_dailies.append([str(player.id), cur_date, 0, 0, 0, 0])
            wb.save("AttackonWikia/records.xlsx")
            time_remaining = get_time_remaining(time.gmtime().tm_hour, time.gmtime().tm_min)
            standard_daily = '⬛ Complete 10 standard puzzles! **[+15 Exp]**'
            standard_daily2 = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ - Progress: **0/10**'
            hangman_daily = '⬛ Win 10 Hangman games! **[+15 Exp]**'
            hangman_daily2 = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ - Progress: **0/10**'
            image_daily = '⬛ Guess 10 images correctly! **[+15 Exp]**'
            image_daily2 = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ - Progress: **0/10**'
            challenge_daily = '⬛ Complete a challenge! **[+25 Exp]**'
            challenge_daily2 = '￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵ ￶￵￶￵ - Progress: **0/1**'

        cur_dailies = discord.Embed(colour = 0xC0C0C0)
        cur_dailies.set_author(name = 'Today\'s Dailies ' + time_remaining, icon_url = player.avatar_url)
        cur_dailies.set_thumbnail(url = player.avatar_url)

        cur_dailies.add_field(name = standard_daily, value = standard_daily2)
        cur_dailies.add_field(name = hangman_daily, value = hangman_daily2)
        cur_dailies.add_field(name = image_daily, value = image_daily2)
        cur_dailies.add_field(name = challenge_daily, value = challenge_daily2)
        return cur_dailies

    def get_stats(self):
        # Returns info of game history in an embed
        wb = load_workbook("AttackonWikia/records.xlsx")
        game_data = wb['Overall']
        game_stats = discord.Embed(title = '📖 Attack on Wikia Stats', colour = 0xC0C0C0)

        total_qns = game_data['A2'].value
        correct_qns = str(game_data['B2'].value)
        if total_qns > 0:
            correct_qns += ' (' + str(round(100 * game_data['B2'].value / total_qns, 1)) + '%)'

        hangman_games = game_data['F2'].value
        hangman_won_games = str(game_data['G2'].value)
        if hangman_games > 0:
            hangman_won_games += ' (' + str(round(100 * game_data['G2'].value / hangman_games, 1)) + '%)'
        
        images_generated = game_data['H2'].value
        images_correct = str(game_data['I2'].value)
        if images_generated > 0:
            images_correct += ' (' + str(round(100 * game_data['I2'].value / images_generated, 1)) + '%)'

        challenge_qns = game_data['D2'].value
        c_correct_qns = str(game_data['E2'].value)
        if challenge_qns > 0:
            c_correct_qns += ' (' + str(round(100 * game_data['E2'].value / challenge_qns, 1)) + '%)'
        
        game_stats.add_field(name = 'Questions asked', value = str(total_qns))
        game_stats.add_field(name = 'Questions answered correctly', value = correct_qns)
        game_stats.add_field(name = 'Total Hangman games played', value = str(hangman_games))
        game_stats.add_field(name = 'Total Hangman games won', value = hangman_won_games)
        game_stats.add_field(name = 'Number of images generated', value = str(images_generated))
        game_stats.add_field(name = 'Images guessed correctly', value = images_correct)
        game_stats.add_field(name = 'Challenges played', value = str(game_data['C2'].value), inline = False)
        game_stats.add_field(name = 'Challenge questions asked', value = str(challenge_qns))
        game_stats.add_field(name = 'Challenge questions answered correctly', value = c_correct_qns)
        
        return game_stats

    def get_leaderboard(self, server, page=1):
        # Returns the names of the top players, with 10 per page, in an embed
        wb = load_workbook("AttackonWikia/records.xlsx")
        player_data = wb['Players']
        
        player_rankings = {} # {player id: sr, ...}
        # Put all players into a dictionary
        for line in player_data:
            player_rankings[line[0].value] = line[1].value

        server_users = server.members
        server_players = []
        for user in server_users:
            if str(user.id) in player_rankings:
                server_players.append([user.mention, player_rankings[str(user.id)]])
        
        # Sort by Exp from biggest to smallest
        server_players.sort(key = lambda x: x[1], reverse = True)

        # Top 10x players
        try: 
            page_no = int(page)
        except:
            page_no = 1

        if page_no < 1:
            page_no = 1

        # Normalize to number of server users
        page_no = min(math.ceil(len(server_players) / 10), page_no)
        
        # Put names into embed
        all_names = ''
        all_levels_exp = ''
        for rank in range((page_no - 1) * 10 + 1, min(len(server_players) + 1, page_no * 10 + 1)):
            all_names += str(rank) + '. ' + server_players[rank-1][0] + '\n'
            for level in self.levelling_system:
                if server_players[rank-1][1] < level[1]:
                    next_level_exp = level[1]
                    break
                player_level = level[0]
            all_levels_exp += '**' + str(player_level) + '**' + ' (' + str(server_players[rank-1][1])
            if player_level < len(self.levelling_system):
                all_levels_exp += '/' + str(next_level_exp)
            all_levels_exp += ')\n'

        leaderboard = discord.Embed(title = 'Leaderboard for Attack on Wikia', colour=0xC0C0C0)
        leaderboard.add_field(name = 'Max Level', value = str(len(self.levelling_system)), inline = False)
        leaderboard.add_field(name = 'Rank/Name', value = all_names)
        leaderboard.add_field(name = 'Level/Exp', value = all_levels_exp)
        return leaderboard

    def get_commands(self):
        # Returns the list of commands
        commands_list = discord.Embed(title = 'List of commands for Attack on Wikia', colour = 0xC0C0C0)
        commands_list.add_field(name = '~new', value = 'Starts a new puzzle in normal mode.')
        commands_list.add_field(name = '~clue', value = 'Use during an active puzzle in normal mode to obtain more information about the wiki page. Maximum of 5 clues per puzzle.')
        commands_list.add_field(name = '~hangman', value = 'Starts a new Hangman game. Type ~hangman <@person> to challenge someone to a Hangman game.')
        commands_list.add_field(name = '~image', value = 'Starts image mode.')
        commands_list.add_field(name = '~answer', value = 'Reveals the answer of an active puzzle or Hangman game')
        commands_list.add_field(name = '~challenge <@person>', value = 'Sends a challenge to another person, or accepts an incoming challenge.')
        commands_list.add_field(name = '~reset', value = 'Resets the game.')
        commands_list.add_field(name = '~profile <@person>', value = 'Checks the profile of a given user. Use just ~profile to check your own profile.')
        commands_list.add_field(name = '~badges <@person>', value = 'Checks the badges a given user has. Use just ~badges to check your own badges.')
        commands_list.add_field(name = '~dailies', value = 'Checks the progress of your daily rewards.')
        commands_list.add_field(name = '~gamestats', value = 'Shows the overall game statistics.')
        commands_list.add_field(name = '~leaderboard/~lb <page>', value = 'Shows the leaderboard, listing the top 10 players on the server. Add a number behind to see subsequent pages (e.g. ~lb 2).')
        return commands_list